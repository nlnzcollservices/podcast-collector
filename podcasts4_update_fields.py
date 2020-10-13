import os
import io
import requests
import hashlib
import gspread
from pymarc import parse_xml_to_array,record_to_xml, Field 
import time
from bs4 import BeautifulSoup
from datetime import datetime as dt
from settings import logging, file_folder, template_folder, pr_key, sb_key, logging, start_xml, end_xml
from openpyxl import load_workbook
from alma_tools import AlmaTools
from database_handler import DbHandler


class Manage_fields():



	def __init__(self, key, mms_id_list):

		self.f347 = None
		self.f500 = None
		self.f856 = None
		self.f942 = None
		self.key = key
		self.alma_key = None
		self.mms_id = None
		self.mms_id_list = mms_id_list
		self.flag = False

		
			
	def removing_dup_fields_add_942(self, field_num):

		"""
		Using pymarc object for finding and removing particular dupped field
		Parameters:
			field_num(str) - number of field

		"""
		
		fields = self.record.get_fields(field_num)

		if len(fields) != 0:
			field_dict = {}
			for ind in range(len(fields)):
				if fields[ind].value() not in field_dict:
					field_dict[fields[ind].value()] = [fields[ind]]
				else:
					field_dict[fields[ind].value()] += [fields[ind]]
					self.flag = True

			if not  self.flag:
				pass
			if self.flag:
				logging.info("needs to be removed")
				self.record.remove_fields(field_num)
				for el in field_dict.keys():
					self.record.add_ordered_field(field_dict[el][0])
				for el in self.record:
					logging.info(el.tag)
					logging.info(el.value().encode("utf-8"))
		else:
			if field_num == "942":
				date_942 = 	(str(dt.now().strftime( '%Y-%m')))
				f942 = Field(tag = '942', indicators = ["",""], subfields = ['a', 'nznb {}'.format(date_942)])
				self.record.add_ordered_field(f942)
						



	def parsing_bib_xml(self):

		""""Converts bib xml to pymarc. Looking for duplicates in self.field_list. Runs removing_dup_fields_add_942 """

		self.record = parse_xml_to_array(io.StringIO(self.bib_data))[0]
		self.field_list = ["347", "500", "856", "942"]
		for field_num in self.field_list:
			self.removing_dup_fields_add_942(field_num)
		self.bib_data = record_to_xml(self.record)
		self.bib_data = start_xml + bib_data +end_xml
	def cleaning_routine(self):

		"""Running routine for modification of bib record"""

		sb_mms = None
		if self.key=="sb":

			sb_mms = "9918602951502836"

		self.f347 = None
		self.f500 = None
		self.f856 = None
		self.f942 = None
		self.mms_id = None
		self.bib_data = None
		for mms in self.mms_id_list:
				self.mms_id = mms
				self.flag_dup = False
				self.flag_no_942 = False
				my_rec = AlmaTools(self.key)
				my_rec.get_bib(mms)
				self.bib_data = my_rec.xml_data
				self.parsing_bib_xml()
				if not self.flag:
					my_rec.update_bib(mms, self.bib_data)
					# print(my_rec.status_code)
					# print(my_rec.xml_data)
					my_db =DbHAndler()
					my_db.db_update_updated(self.mms_id)





def main():


	mms_list = []

	workook_path = r"D:\\dup_942.xlsx"
	if os.path.exists(workook_path):

		wb = load_workbook(workook_path)
		#Enter name of the working sheet bellow
		ws= wb.get_sheet_by_name("results")
		#if now headers min_row =1
		for row in ws.iter_rows(min_row=2):

		#depending on where mms id is row[3] should be changed to number of column started from 0.
			mms = row[21].value
			mms_list.append(row[21].value)

		my_rec = Manage_fields("prod",mms_list)
		my_rec.cleaning_routine()	

		


	pass
if __name__ == '__main__':

	
	main()