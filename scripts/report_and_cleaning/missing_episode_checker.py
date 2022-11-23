
import re
import io
import os
import json
from fuzzywuzzy import fuzz
# from settings_prod import  report_folder, assets_folder
from datetime import datetime as dt
import feedparser
import dateparser
from word2number import w2n
from podcast_dict import podcasts_dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
import openpyxl
import random
import ssl
if hasattr(ssl, '_create_unverified_context'):
    ssl._create_default_https_context = ssl._create_unverified_context




report_folder =r"Y:\ndha\pre-deposit_prod\LD_working\podcasts\log\reports"

time_string = "_"+dt.now().strftime("%m_%Y")
cleaning_report_folder = "cleaning_report"+time_string
cleaning_report_folder_path = os.path.join(report_folder, cleaning_report_folder)
print(cleaning_report_folder_path)
json_file_path = os.path.join(cleaning_report_folder_path, "podcast_cleaning_report.json")
print(json_file_path)

dates_list = []



wb = Workbook()

def clean_ep_title(title):
	title = title.replace("Episode","").replace("Ep","").replace("EP","").replace("podcast","").replace("The Chris and Sam","").replace("|","").replace("-","").replace("Taringa","")
	return title
def clean_pod_title(title):
	title = title.replace("podcast","").replace("Podcast","")
	if ":" in title:
		title = title.split(":")[0]
	if "/"  in title:
		title = title.split("/")[0]
	return title

def  make_spreadsheet():


	thin_border = Border(left=Side(style='thin'), 
	right=Side(style='thin'), 
	top=Side(style='thin'), 
	bottom=Side(style='thin'))
	ws1 = wb["Sheet"]
	ws1.cell(column=1, row=ws1.max_row, value = "PODCASTS")
	my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="FFAC00")
	ws1.cell(column=3, row=ws1.max_row, value = "episodes").fill=my_fill
	ws1.cell(column=3, row=ws1.max_row).border = thin_border
	my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="C4B07B")
	ws1.cell(column=4, row=ws1.max_row, value = "mixed").fill=my_fill
	ws1.cell(column=4, row=ws1.max_row).border = thin_border
	my_fill= openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="95C8F0")
	ws1.cell(column=5, row=ws1.max_row, value = "date").fill=my_fill
	ws1.cell(column=5, row=ws1.max_row).border = thin_border



	date_series_dict = make_inverse_dict()
	for serial_tit in date_series_dict.keys():
		
		title_list = []
		not_used = []
		ser_tit = serial_tit.replace(" ",'_').replace(",","").replace("!","").replace("'","").replace("/","").replace(":","").replace("$","S")
		rss_link = None
		for d_tit in podcasts_dict.keys():
			if title_list ==[]:
				if  fuzz.ratio(clean_pod_title(d_tit) ,  clean_pod_title(serial_tit))>80:
					try:
						rss_link = podcasts_dict[d_tit]["rss_filename"]
						print(rss_link)
						title_list = check_rss( rss_link, None)
					except:
						print("no rss link")
	
		for ti in title_list:
			not_used.append(ti[0])
		try:

			wb.create_sheet(ser_tit)
			ws = wb[ser_tit]
			hexadecimal = ''.join([random.choice('ABCDE') for i in range(6)])
			my_color = openpyxl.styles.colors.Color(rgb=hexadecimal)
			ws.sheet_properties.tabColor = my_color
			count_new = 0
			fuz_flag = False
			my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
			thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
			thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))
			ws.cell(column=1, row=ws.max_row, value=serial_tit).fill=my_fill
			link = "#Sheet!A1"
			ws.cell(column=1, row=ws.max_row).hyperlink = link
			ws.cell(column=1, row=ws.max_row).border = thick_border
			ws.cell(column=2, row=ws.max_row, value="Number/Date").fill=my_fill
			ws.cell(column=2, row=ws.max_row).border = thin_border
			ws.cell(column=3, row=ws.max_row, value="MMS id").fill=my_fill
			ws.cell(column=3, row=ws.max_row).border = thin_border
			ws.cell(column=4, row=ws.max_row, value="yyyymmdd").fill=my_fill
			ws.cell(column=4, row=ws.max_row).border = thin_border
			ws.cell(column=5, row=ws.max_row, value="Matched").fill=my_fill
			ws.cell(column=5, row=ws.max_row).border = thin_border
			ws.cell(column=6, row=ws.max_row, value="Date").fill=my_fill
			ws.cell(column=6, row=ws.max_row).border = thin_border
			ws.cell(column=7, row=ws.max_row, value="").fill=my_fill
			ws.cell(column=7, row=ws.max_row).border = thick_border
			ws.cell(column=8, row=ws.max_row, value=rss_link).fill=my_fill
			ws.cell(column=8, row=ws.max_row).border = thick_border
			ws.cell(column=9, row=ws.max_row, value="Date").fill=my_fill
			ws.cell(column=9, row=ws.max_row).border = thick_border

			
			ws.column_dimensions["A"].width = 20
			row_count =int(ws.max_row)+1
			all_count = 0
			date_count =0
			for el in date_series_dict[serial_tit].keys():
				all_count +=1
				ws.cell(column=1, row=row_count, value=el).border = thin_border
				ws.cell(column=2, row=row_count, value = date_series_dict[serial_tit][el][0]).border = thin_border
				ws.cell(column=3, row=row_count, value = date_series_dict[serial_tit][el][1]).border = thin_border
				try:
					# print(date_series_dict[serial_tit][el][0])
					date_changed = dateparser.parse(date_series_dict[serial_tit][el][0]).strftime("%Y%m%d")
					# print(date_changed)
					if date_changed:
						ws.cell(column=4, row=row_count, value = date_changed).border = thin_border
						date_count +=1
				except:
					pass
				row_count+=1
				if title_list != []:
					for r_tit in title_list:
						if  fuzz.ratio(clean_ep_title(r_tit[0]) ,  clean_ep_title(el))>70:
							ws.cell(column=5, row=row_count-1, value = r_tit[0]).border = thin_border
							ws.cell(column=6, row=row_count-1, value = r_tit[1]).border = thin_border

							for index, nu in enumerate(not_used):
								if nu==r_tit[0]:
									not_used.pop(index)


			for i,left in enumerate(not_used):
				ws.cell(column=8, row=i+2, value = left).border = thick_border
				for tt in title_list:
					if tt[0]==left:
						ws.cell(column=9, row=i+2, value = tt[1]).border = thick_border



			ws1 = wb["Sheet"]
			link = "#{}!A1".format(ser_tit)
			if date_count/all_count < 0.1:#episodes
				my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="FFAC00")
			elif date_count/all_count >=0.1 and date_count/all_count<0.9:#ep and date
				my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="C4B07B")
			elif date_count/all_count>=0.9:
				my_fill= openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="95C8F0")
			ws1.cell(column=1, row=ws1.max_row+1, value = serial_tit).fill=my_fill#.hyperlink = link
			ws1.cell(column=1, row= ws1.max_row).hyperlink = link
			ws1.column_dimensions["A"].width = 30

			# 	for el in d["entries"]:
			# 		print("here8")
			# 		new_date = dateparser.parse(el["published"]).strftime("%B %d %Y")
			# 		print(fuzz.ratio(el["title"] ,  title))
			# 		print("here9")
			# 		if  fuzz.ratio(el["title"] ,  title)>70:
			# 			print("here10")
			# 			ws.cell(column=3, row=row_count, value = el["title"])
			# 			ws.cell(column=4, row=row_count, value = new_date)
			# 			rss_dict[series_title][0].append(el["title"])
					
			# 	row_count+=1
		except Exception as e:
			print(str(e))
			print("wrongwrong")
			print(serial_tit)
	wb.save("test.xlsx")
	# json.dump(rss_dict,"test.json")
	# wb = load_workbook("test.xlsx")
	# rss_doct=json.load("test.json")
	# for  tit   in  rss_dict.keys():
	# 	ws = wb[tit]
	# 	max_row = ws.max_row
	# 	for c in 10:
	# 		ws.cell(column=c, row=max_row, value = "-------").fill = my_fill
	# 		for el  in tit[1]["entries"]:

	# 			if not el["title"] in tit[0]:
	# 				ws.cell(column=1, row=ws.max_row, value = el["title"])
	# 				ws.cell(column=2, row=ws.max_row, value = dateparser.parse(el["published"]).strftime("%B %d %Y"))
	# wb.save("test.xlsx")



					


				# and not "S2" in title_dict[title][mms]["series"] and not "S3" in title_dict[title][mms]["series"] and not "S3" in title_dict[title][mms]["series"]:
#				print(title_dict[title][mms]["series"])

	# 	ws_podc =  wb.create_sheet(title=title)
	# 	header_row_1 = ["", "APP","", "", "PRO","", "", "DA", "", "",  "AVC","", ""]
	# 	header_row_2 = ["Year", "Collections", "Files", "Size", "Collections", "Files", "Size", "Collections", "Files", "Size", "Collections", "Files", "Size"]

	# 	### make the two headings rows
	# 	for i, field in enumerate(header_row_1):
	# 		field = statuses[field]
	# 		ws_podc.cell(column=i+1, row=row_counter, value="{0}".format(field))
	
	# 	ws_collection.merge_cells('B1:D1') 
	# 	ws_collection['B1'].alignment = Alignment(horizontal="center", vertical="center")
	# 	ws_collection['B1'].font = Font(bold=True)
	# 	ws_collection.merge_cells('E1:G1')
	# 	ws_collection['E1'].alignment = Alignment(horizontal="center", vertical="center")
	# 	ws_collection['E1'].font = Font(bold=True)
	# 	ws_collection.merge_cells('H1:J1')
	# 	ws_collection['H1'].alignment = Alignment(horizontal="center", vertical="center")
	# 	ws_collection['H1'].font = Font(bold=True)
	# 	ws_collection.merge_cells('K1:M1')
	# 	ws_collection['K1'].alignment = Alignment(horizontal="center", vertical="center")
	# 	ws_collection['K1'].font = Font(bold=True)
	# 	row_counter += 1
	# 	for i, field in enumerate(header_row_2):
	# 		ws_collection.cell(column=i+1, row=row_counter, value="{0}".format(field)).font = Font(bold=True)
	# 	row_counter += 1


	# 	### populate the per year data
	# 	for y in range(int(min(years[0].keys())), now.year+1):
	# 		ws_collection.cell(column=1, row=row_counter, value="{0}".format(y)).font = Font(bold=True)

	# 		if str(y) in years[0].keys():
	# 			# do line filling
	# 			my_row = do_line_shaping(years[0][str(y)])
	# 			for i, field in enumerate(my_row):
	# 				try: 
	# 					field = int(field)
	# 					ws_collection.cell(column=i+2, row=row_counter, value="{:,}".format(field))
	# 				except ValueError:
	# 					ws_collection.cell(column=i+2, row=row_counter, value="{0}".format(field))

	# 		else:
	# 			ws_collection.cell(column=2, row=row_counter, value="N/A")
	# 		row_counter += 1

	# 	### make the sub total row
	# 	row_counter+=2
	# 	my_row = do_collection_totals(years)
	# 	for i, field in enumerate(my_row):
	# 		try: 
	# 			field = int(field)
	# 			ws_collection.cell(column=i+1, row=row_counter, value="{:,}".format(field))
	# 		except ValueError:
	# 			ws_collection.cell(column=i+1, row=row_counter, value="{0}".format(field))

	# 	ws_collection.cell(column=1, row=row_counter).font = Font(bold=True)
		
	# 	collection_size = 0
	# 	collection_file_count = 0
	# 	collection_folder_count = 0

	# 	for year, data in years[0].items():
	# 		for status, data in data.items():

	# 			total_size += data["size"]
	# 			collection_size += data["size"]

	# 			total_file_count += data["file_count"]
	# 			collection_file_count += data["file_count"]

	# 			total_folder_count += data["folder_count"]
	# 			collection_folder_count += data["folder_count"]
	# 	collection_totals[collection] = [collection_folder_count, collection_file_count, humanize.naturalsize(collection_size)]

	# #### do the summary page
	# collections_keys = list(collection_totals.keys()) 
	# collections_keys.sort()

	# row_counter = 3

	# ws1.cell(column=1, row=row_counter, value="{0}".format("Collection Area")).font = Font(bold=True)
	# ws1.cell(column=2, row=row_counter, value="{0}".format("Collections")).font = Font(bold=True)
	# ws1.cell(column=3, row=row_counter, value="{0}".format("Files")).font = Font(bold=True)
	# ws1.cell(column=4, row=row_counter, value="{0}".format("Size")).font = Font(bold=True)
	# row_counter+= 1


	# ### per collection area
	# for key in collections_keys:
	# 	ws1.cell(column=1, row=row_counter, value="{0}".format(collections[key])).font = Font(bold=True)
	# 	for i, field in enumerate(collection_totals[key]):
	# 		try: 
	# 			field = int(field)
	# 			ws1.cell(column=i+2, row=row_counter, value="{:,}".format(field))
	# 		except ValueError:
	# 			ws1.cell(column=i+2, row=row_counter, value="{0}".format(field))
	# 	row_counter+= 1

	# ### total row
	# row_counter+= 1
	# ws1.cell(column=1, row=row_counter, value="{0}".format("Total")).font = Font(bold=True)
	# ws1.cell(column=2, row=row_counter, value="{0}".format(total_folder_count))
	# ws1.cell(column=3, row=row_counter, value="{:,}".format(total_file_count))
	# ws1.cell(column=4, row=row_counter, value="{0}".format(humanize.naturalsize(total_size)))

	# ### fin
	wb.save("test.xlsx")


def make_inverse_dict():
	date_series_dict =  {}
	series_list = []
	with open(json_file_path, "r") as f:
		title_dict = json.load(f)
	for title  in title_dict.keys():
		my_title = title.rstrip(".")
		for  mms in title_dict[title]:
			series_title = title_dict[title][mms]["series"].split(";")[0].rstrip(" ,")#+"|"+title +"|"+title_dict[title][mms]["series"].split(";")[-1].lstrip(" ")
			# print(series_title)
			title_date =  title_dict[title][mms]["series"].split(";")[-1].lstrip(" ").replace(",","").strip(".")
			series_list.append(series_title)
			if not "BYC" in series_title:
				if not series_title in date_series_dict:
					date_series_dict[series_title]={my_title :[title_date,mms]}
				else:
					date_series_dict[series_title][my_title ] = [title_date,mms]
	# for el in date_series_dict:
	# 	print("#"*50)
	# 	print(el)
	# 	for e in date_series_dict[el]:
	# 		print("-"*50)
	# 		print(e)
	# 		print(date_series_dict[el][e])
	return  date_series_dict

	

def check_for_series(series_title):
	"""This function is gathering dates from cleaning _report.json
	Parameters:
		series_title(str) - title of podcast of interest
	Returns:
		dates_list(list) - list of dates

	"""
	print("Checking from Alma set json report")
	series_list = []
	date_dict = {}
	
	with open(json_file_path, "r") as f:
		title_dict = json.load(f)

	for title  in title_dict.keys():
		for  mms in title_dict[title]:
			series_list.append(title_dict[title][mms]["series"].split(";")[0]+"|"+title +"|"+title_dict[title][mms]["series"].split(";")[-1].lstrip(" "))
			if series_title in title_dict[title][mms]["series"]:
				
				dates_list.append(title_dict[title][mms]["series"].split(";")[-1].lstrip(" ").replace(",","").strip("."))
				date_dict[title] = title_dict[title][mms]["series"].split(";")[-1].lstrip(" ").replace(",","").strip(".")
	print("Episodes in Alma")
	for el in dates_list:
		print(el)
	print("*"*50)
	return dates_list, series_list, date_dict


def check_missing_ep(series_title, count=0):

	ep_dict = {}
	print(series_title)

	with open(json_file_path, "r") as f:
		title_dict = json.load(f)
	for title  in title_dict.keys():
		for  mms in title_dict[title]:
			if series_title in title_dict[title][mms]["series"]:# and not "S2" in title_dict[title][mms]["series"] and not "S3" in title_dict[title][mms]["series"] and not "S3" in title_dict[title][mms]["series"]:
#				print(title_dict[title][mms]["series"])
				ep_number =	title_dict[title][mms]["series"].split(";")[-1].strip("[]")
				if "//" in ep_number:
					ep_number = ep_number.split("//")[0]
				ep_number = ep_number.lstrip(" ").strip(".").rstrip(" ").split(" ")[-1].strip("#")
				ep_number = ep_number.strip("EpP").strip(".")
				if ep_number.isdigit() and int(ep_number)>1990:
					ep_number = title.replace(".","").split(" ")[-1]
				if not ep_number.isdigit():
					ep_number_string = title_dict[title][mms]["series"].split(";")[0]
					ep_number_list = re.findall(r"\d{3}", ep_number_string)
					if len(ep_number_list)>0:
						ep_number = ep_number[0]
				if ":" in ep_number:
					ep_number = ep_number.split(":")[-1].strip("EP")
				try:
					ep_number = w2n.word_to_num(ep_number)
					#print(ep_number)
					ep_number =str(ep_number).zfill(3)
					#print(ep_number)

				except:
					
					pass


				ep_dict[title] = ep_number
				
	ep_dict = dict(sorted(ep_dict.items(), key=lambda item: item[1]))


	for title in ep_dict:


		num = ep_dict[title]
		count+=1
		#print(num)
		# print(title)
		#print(count)
		#print("#"*50)
		if num.isdigit() and int(num)<1200:
			try:

				if int(num) != count:


					print("-"*50)
					print("-"*50)

					difference = int(num)-count

					for i in list(range(difference)):

						print(count+i)

					#print(count)
					print("-"*50)
					count+=difference
			except:
				print("Non standard: ", num ,title)
		else:
			print("Non standard: ", num, title)
	return ep_dict



def check_rss(rss_link,dates_list = None):

	"""This function is for those podcasts which have dates in 490
	Parameters:
		dates_list(list) - list of dates of episodes in Alma
		rss_link(str) - rss link


	"""

	print("Checking rss feed for ",rss_link)
	#March 30 2020
	d = feedparser.parse(rss_link)
	# print("here777777")
	rss_list = []
	title_list  = []
	for i,elem in enumerate(d["entries"]):
		# print(elem)
		# print(elem["title"])
		new_date = dateparser.parse(elem["published"]).strftime("%B %d %Y")
		rss_list.append(new_date)
		title_list.append([elem["title"],new_date])
		count_missing = 0
		
		if dates_list:
			if not new_date in dates_list:
				
				print(elem["title"], new_date, " - missing")
				count_missing+=1
	if count_missing == 0:
		print("Collected everything from rss")
	else:
		print(count, " - missing")

	return title_list

if __name__ == '__main__':
	rss_link = None
	#series_title = "Coronavirus podcast" 
	#rss_link  = r"https://www.rnz.co.nz/podcasts/the-coronavirus-podcast.rss"
	#series_title = "Crave!"  #72,75,76,77,84,87,88,89 -different  workflow
	series_title  = "Actually interesting"
	#series_title = "Taringa" #Done(134,144,145,185,200)
	#series_title = "Dirt Church radio" #Done 162
	#series_title = "Better off read" # from 72-81,85 mixed title - need help
	#series_title = "BYC" #before 100 and 157 and question about current harvesting
	#series_title = "Retrogasmic" #ok
	#series_title = "Ciaran McMeeken" #strange rss to check possibli ep 10 missed
	#series_title = "Stirring the pot"#mix difficult to check probably one missing
	#series_title = "NZSA oral history podcast" #mix difficult to say
	# series_title = "Never repeats" #1-29 missing cannot check as rss not working
	#series_title = "Rubbish trip" ok 
	#series_title = "Chris" #mix
	#series_title = "History of Aotearoa New Zealand podcast"# check
	#series_title = "Advanced analytics" #S1, S2, S3  #Very messy rss feed
	#series_title = "Back to the disc-player podcast" ok
	#series_title = "Stupid questions for scientists" ok
	#series_title = "Snacks and chats" #ok
	#series_title = "verbal highs" #probably ok
	#series_title = "NZ tech podcast" #mix difficult to check
	#series_title = "Property academy podcast" #need to be collected from 868
	#series_title = "CIRCUIT cast" #1,2,3 missing from rss Done,10,87-90
	#series_title = "UC Science radio" #ok
	#series_title = "Dietary requirements"#mix difficult to say
	#series_title = "Love this podcast"# Done3-61,64,70,72,73
	#series_title = "your day job" ok
	#series_title = "Animal matters"# Done 32
	#series_title = "Retrogasmic"# ok
	#series_title = "Selfie reflective"ok
	#series_title = "die in bed" #ok
	# series_title = "verbal highs"#ok
	# series_title = "Lip podcast"#ok
	#series_title = "Windows on dementia"# Done2 +  all new
	# series_title = "Classic NBL podcast"# ok check if seased
	# series_title =  "New Zealand brewer podcast" #mixed difficult seems ok
	# series_title = "Girls on top" #ok
	# series_title = "Angus Dunn podcast" #ok
	# series_title = "HP business class" #seems ok
	# series_title = "Kiwi country" #seems ok
	# series_title = "Queenstown life" #ok
	# series_title = "Few good men"#ok
	# series_title = "EPIC podcast"#ok check if finished
	# series_title = "Mums in cars"#ok
	# series_title = "Mud & blood" #seems to be ok difficult to say
	# series_title = "History of Aotearoa New Zealand podcast"#seems ok
	# series_title = "Seeing"
	# series_title = "76 small rooms"#Episode 039
	# rss_link = "https://feeds.soundcloud.com/users/soundcloud:users:151008528/sounds.rss"
	# series_title = "Rubbish trip" #seems ok
	# series_title = "Baboon yodel" #seems ok 110 not present first 40 missing
	# series_title = "Stag roar" #difficult to say very mixed
	# series_title = "Kiwi birth tales" #seems to be ok mixed
	# series_title = "Cult popture"
	# series_title =  "Gone by lunchtime"
	# series_title = "How to save the world"
	#series_title = "Taxpayer talk" #last was august 2021
	# series_title = "B-side stories" #in spreadsheet
	# series_title = "Boners of the heart podcast"#ok
	# series_title = "Book bubble" #ok
	# series_title = "Bosses in lockdown"# ok
	# series_title = "Bosses rebuilding"#not ok but could be related with tile change
	# series_title = "Brazen"#ok
	# series_title = "Breeder's digest" 
	# series_title = "Business is boring" #possibly not before 23 november 2018 !!!!!!!!!!!!!! Check carefully!!!!!
	# series_title = "Carat chats"# missed"The Fourth Estate is dead  October14 2019"
	# rss_link = "https://www.spreaker.com/show/4105065/episodes/feed"
# 	series_title = "Conversations that count" #ok new missed
# 	series_title = "Cooking the books"#seems to be ok new missed
# 	series_title = "Coronavirus podcast"#Back to School April 09 2020 was inverstigated by Rhonda
# 	series_title = "DOC sounds of science podcast" #mixed 0k
	#series_title = "Property academy"
	#series_title = "Taxpayer talk"

# 	series_title = "Democracy Project" #probably missed
# 	"""Welfare reform and lowering the voting age – Democracy Project review September 21 2020
# Labour tax and dental care policies – Democracy Project review September 13 2020
# Wayne Mapp – NZ’s foreign policy & defence strategy September 11 2020
# Green Party private school controversy – Democracy Project weekly review September 07 2020"""
# 	#series_title = "Dietary requirements" #mixed need to be checked not in  alma 2019,1018 and some 2020
# 	#series_title = "Fedtalks" #???missed
# 	series_title = "Fold"#???missed
# 	series_title = "Frickin Dangerous Bro show"#??? missing??
# 	series_title = "Going viral" #Not  in dictionary
# 	series_title = "Gone by lunchtime" # ok new to collect
	# series_title = "Goodfellow podcast" #not in dictionary new to collect
	# rss_link = "https://www.goodfellowunit.org/podcast/rss"
	#series_title = "Hosting" #Day One December 12 2016
	#series_title = "How to save the world" #ok
	#series_title = "Human-robot interaction podcast" # ok new missed
	# series_title = "Indigenous urbanism" #not in dict ok
	# rss_link = "http://indigenousurbanism.net/rss"
	#series_title = "Kiwi yarns"# ok not in dict
	#rss_link = "https://plums-dalmatian-yzw6.squarespace.com/kiwi-yarns?format=rss"
	#series_title = "Locals only" #mixed need help
	#series_title = "Lunch money"#ok new to collect
	#serial_title = "Male gayz"  #ok new to collect
	# series_title = "Maxim Institute podcast" #new to collect
	# series_title = "A moment in crime" #seems ok
	# series_title = "Motherness" # ok not in dict
	# rss_link = "https://feeds.buzzsprout.com/853762.rss"
	#series_title = "Mud & blood" #mixed  difficult to say
	#series_title = "NZ history"#ok new to collect
	# series_title = "NZSA oral history podcast" #not in dict mixed 
	# rss_link = "https://feeds.soundcloud.com/users/soundcloud:users:449483649/sounds.rss"
	#series_title = "New Zealand Initiative" #Are we making NCEA credits too easy for kids to get?#Podcast: The pulse of the property market July 13 2021
	#series_title = "New Zealand business tips"#ok
	# series_title = "Offspin"#The Offspin podcast: Jimmy Neesham on life, loss and that runout August 26 2019
	# rss_link = "https://rss.acast.com/the-offspin"
	#series_title = "On the rag" #On the Rag: Scrambled brains, bad cartoons and the Freddy Krueger in the room October 31 2019
	#series_title = "Paige's space"# ok
	#series_title = "Papercuts"#seems to be ok
	# series_title = "Phoenix City"#seems to be ok mixed not in dict
	# rss_link = "http://thekidsareallwhite.podomatic.com/archive/rss2.xml"
	#series_title = "Pod on the couch"#"???   Missing
	# series_title = "Polidic" #ok
	# rss_link = "https://www.omnycontent.com/d/playlist/6e2a797b-0cc4-4c0b-a44d-a51e0019bc3c/f5c796b6-19ca-41d0-8bff-a68c00df11ad/193caf0a-87f4-4441-934d-a68c00e44c3e/podcast.rss"
	# series_title = "Politics in full" #ok
	# rss_link ="https://feeds.simplecast.com/mHiTLFSx"
	#series_title = "Purpose fuelled performance" #Check DONE(5-12,17,18,23) To manuall collect - 36
	# series_title = "Real pod"
	# rss_link ="https://rss.acast.com/the-real-pod" #2019,2020 backlog check as it is mixed
	#series_title = "Seeds"#backlog before 25 Dec 2018
	#series_title = "Shit show"#ok collect new
	# series_title = "Sick leave podcast"# ok not in dict 
	# rss_link = "https://feed.podbean.com/shaunkanderson/feed.xml "
	#series_title = "Stag roar"# ok collect new
	#series_title = "Stirring the pot" #mixed difficult to check
	# series_title = "The front page"# ok some on spreadsheet
	# rss_link = "https://www.spreaker.com/show/3208188/episodes/feed"
	# series_title = "Tick tick"# some stuck could not check
	# rss_link = ""
	#series_title = "Top writers radio show"#seems ok
	#series_title = "True crime New Zealand" #Case 23: The Aramoana Massacre (PART I) November 04 2020 #Case 1: Parker-Hulme Murder (Part 2) June 21 2019
	#series_title = "Victoria University of Wellington podcast" #ok new to collect
	#series_title = "Walk out boys"#????Little Empire Lolly Scramble January 12 2017
	# series_title = "Watercooler"
	# rss_link =  "https://www.omnycontent.com/d/playlist/6e2a797b-0cc4-4c0b-a44d-a51e0019bc3c/732e5172-1388-4943-a6e8-aa0e0034844c/2775fe95-d483-495b-b807-aa0e00348451/podcast.rss"
	"""
	Akld #43 FEAST November 26 2017
Akld #42 Raining Cats & Dogs November 20 2017
Akld #41 Bullsh*t November 19 2017
Akld #40 Growing Pains November 19 2017
Akld #39 Metamorphosis June 30 2017
Akld #37 In The Flesh June 11 2017
Akld #36 Flavours of Awkward March 28 2017
Akld #35 Lighting The Fire March 28 2017
Wgtn #13: Was It Something I Said? October 23 2016
Wgtn #12: First Time/Last Time October 16 2016
Wgtn #8: Pride and Humiliation September 24 2016
Auckland Issue: Neighbours January 16 2016
Akld #53 Secret Lives of Drag March 12 2019
Akld #53 Secret Lives of Drag March 12 2019
	"""

	#series_title = "What comes after what comes next"#ok
	# series_title = "Worst idea of all time"#HOSTING 3: 07 The Prestige March 17 2020, Little Empire Lolly Scramble January 12 2017
	# rss_link = "https://rss.acast.com/worstideaofalltime"
	# series_title  ="thehappy$aver.com" # ok collect new
	# rss_link = "https://www.thehappysaver.com/podcast?format=rss"




	if not rss_link:
		rss_link = podcasts_dict[series_title]["rss_filename"]



	


	#rss_link =r"https://feeds.transistor.fm/lets-talk-dementia"


	# dates_list, series_list = check_for_series(series_title)
	# print("#"*50)
	# print("In Alma:")
	# print("#"*50)
	# print("Date")
	# for date in dates_list:
	# 	print(date)# # print(dates_list)
	# print("#"*50)
	# # print("Serial titles")
	# # print("#"*50)
	# # for tit in sorted(series_list):
	# # 	print(tit)
	# print("#"*50)
	# print("In RSS possibly not in Alma")


	# print("#"*50)
	#get Alma episodes from report
	dates_list, series_list, date_dict = check_for_series(series_title)
	#check if rss episode not in Alma report
	check_rss(rss_link, dates_list)
	# print("#"*50)



	# initial_count = 0
	# check_missing_ep(series_title, initial_count)



	#make_spreadsheet()