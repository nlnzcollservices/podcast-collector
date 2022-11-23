from alma_tools import AlmaTools
import re
import io
import os
from pymarc import parse_xml_to_array,record_to_xml, Field 
from openpyxl import load_workbook
import json
from report_viz_tool import json_viz
from settings_prod import  report_folder, assets_folder
from datetime import datetime as dt

set_id = "12558695740002836" #this is full podcasts set id
#it is very large, it could be filtered by date
time_string = "_"+dt.now().strftime("%m_%Y")
cleaning_report_folder = "cleaning_report"+time_string
cleaning_report_folder_path = os.path.join(report_folder, cleaning_report_folder)
if not os.path.isdir(cleaning_report_folder_path):
	os.mkdir(cleaning_report_folder_path)
json_file_path= os.path.join(cleaning_report_folder_path, "podcast_cleaning_report.json")
json_second_path = os.path.join(cleaning_report_folder_path,"podcast_possible_dups.json")
json_third_path = os.path.join(cleaning_report_folder_path,"podcasts_most_problematic.json")
no_rep_no_hold_file = os.path.join(cleaning_report_folder_path,"no_reps_no_holdings.txt")
no_rep_hold_file = os.path.join(cleaning_report_folder_path,"no_reps_holdings.txt")
no_hold_rep_file = os.path.join(cleaning_report_folder_path,"no_hold_rep_file.txt")

filepath = os.path.join(assets_folder,"results.xlsx")
report_cleaning_temp = os.path.join(cleaning_report_folder_path, "report_cleaning_temp_count.txt")
wb = load_workbook(filepath)
ws = wb["results"]

my_api = AlmaTools("prod")
file_name = "podcasts_automated.txt"

def target_set(cleaning):
	if cleaning:
		os.remove(json_file_path)
		try:
			os.remove(report_cleaning_temp)
		except Exception as e:
			print(str(e))
	
	try:
		with open(json_file_path, 'r') as f:
			title_dict = json.load(f)
	except Exception as e:
		print(str(e))
		title_dict = {}
	print(title_dict)
	try:
		with open (report_cleaning_temp,"r") as f:
			my_number=int(f.read())
	except:
		my_number = 0
	print(my_number)
	count_mmss = 0
	my_api.get_set_members(set_id,{"limit":"100"})
	print(my_api.xml_response_data)
	total_count = re.findall(r'<members total_record_count="(.*?)">',my_api.xml_response_data)[0]
	for i in range((int(total_count)//100)+2):		
		my_api.get_set_members(set_id,{"limit":"100","offset":100*i})
		mms_ids = re.findall(r"<id>(.*?)</id>",my_api.xml_response_data)
		for mms in mms_ids:
			count_mmss+=1
			if count_mmss > my_number or my_number==0:

				my_number+=1
				print(my_number)
				with open(report_cleaning_temp,"w") as f:
					f.write(str(my_number))
				
				

				try:
					my_api.get_bib(mms)
				except:
					my_api.get_bib(mms)
				record = parse_xml_to_array(io.StringIO(my_api.xml_response_data))[0]

				mms_id = record["001"].value()
				title = record["245"].value()
				series = str(record["490"]["a"])+"; "+str(record["490"]["v"])+"."
				rep_ids = []
				holdings = {}
				items_list = []

				reps = {}
				ies = {}
				print(title)
				my_api.get_holdings(mms, {"limit":"100"})
				try:

					num_hold = re.findall(r'<holdings total_record_count="(.*?)">',my_api.xml_response_data)[0]
					for i in range(len(num_hold)):
						hold_id = re.findall(r'holdings/(.*?)"',my_api.xml_response_data)[0]
						my_api.get_items(mms,hold_id,{"limit":"100"})
						items_list = []
						try:
							num_items = re.findall(r'<items total_record_count="(.*?)">',my_api.xml_response_data)[0]
							for i in range(len(num_hold)):	
								item_pid = re.findall(r'<pid>(.*?)</pid>',my_api.xml_response_data)[0]
								items_list.append(item_pid)
							holdings[hold_id] = items_list

						except Exception as e:
							if '<items total_record_count="0"/>' in my_api.xml_response_data:
					 			num_items = 0

				except Exception as e:

					if '<holdings total_record_count="0"/>' in my_api.xml_response_data:
					 	num_hold = 0
					 	num_items = 0
				print(holdings)
				print(num_hold)
				print(num_items)
				my_api.get_representations(mms,{"limit":"100"})


				try:
					num_rep = re.findall(r'<representations total_record_count="(.*?)">',my_api.xml_response_data)[0]
					for i in range(len(num_rep)):
							rep_id = re.findall(r"<id>(.*?)</id>",my_api.xml_response_data)[0]
							rep_ids.append(rep_id)
							ie = re.findall(r"pubam:(.*?)</",my_api.xml_response_data)[0]
							label = re.findall(r'<label>(.*?)</label', my_api.xml_response_data)[0]
							print(ie)
							print(label)
							ies[ie] = label
					print(ies)
					print(num_rep)


				except:
				 	if '<representations total_record_count="0"/>' in my_api.xml_response_data:
				 	 	num_rep = 0


				if title in title_dict:
					title_dict[title][mms] = {"ie":ies, "holdings":holdings, "series":series}
					 
				else:
				 	mms_data = {"ie":ies, "holdings":holdings, "series":series}
				 	title_data= {mms:mms_data}
				 	title_dict[title] = title_data
				#print(title_dict)
				with open(json_file_path, 'w') as f:
					json.dump(title_dict, f)



def parse_spreadsheet(cleaning):
	if cleaning:
		os.remove(json_file_path)
		os.remove(report_cleaning_temp)
	
	try:
		with open(json_file_path, 'r') as f:
			title_dict = json.load(f)
	except Exception as e:
		print(str(e))
		title_dict = {}
	print(title_dict)
	try:
		with open (report_cleaning_temp,"r") as f:
			my_number=int(f.read())
	except:
		my_number = 2
	for i, row in enumerate(ws.iter_rows(min_row=my_number)):
		with open("report_cleaning_temp_count.txt","w") as f:
			f.write(str(i+my_number))
		rep_ids = []
		holdings = {}
		items_list = []
		title = row[4].value
		mms = row[26].value
		series = row[3].value
		reps = {}
		ies = {}
		print(title)
		my_api.get_holdings(mms, {"limit":"100"})
		try:

			num_hold = re.findall(r'<holdings total_record_count="(.*?)">',my_api.xml_response_data)[0]
			for i in range(len(num_hold)):
				hold_id = re.findall(r'holdings/(.*?)"',my_api.xml_response_data)[0]
				my_api.get_items(mms,hold_id,{"limit":"100"})
				items_list = []
				try:
					num_items = re.findall(r'<items total_record_count="(.*?)">',my_api.xml_response_data)[0]
					for i in range(len(num_hold)):	
						item_pid = re.findall(r'<pid>(.*?)</pid>',my_api.xml_response_data)[0]
						items_list.append(item_pid)
					holdings[hold_id] = items_list

				except Exception as e:
					if '<items total_record_count="0"/>' in my_api.xml_response_data:
			 			num_items = 0

		except Exception as e:

			if '<holdings total_record_count="0"/>' in my_api.xml_response_data:
			 	num_hold = 0
		print(holdings)
		print(num_hold)
		print(num_items)
		my_api.get_representations(mms,{"limit":"100"})


		try:
			num_rep = re.findall(r'<representations total_record_count="(.*?)">',my_api.xml_response_data)[0]
			for i in range(len(num_rep)):
					rep_id = re.findall(r"<id>(.*?)</id>",my_api.xml_response_data)[0]
					rep_ids.append(rep_id)
					ie = re.findall(r"pubam:(.*?)</",my_api.xml_response_data)[0]
					label = re.findall(r'<label>(.*?)</label', my_api.xml_response_data)[0]
					print(ie)
					print(label)
					ies[ie] = label
			print(ies)
			print(num_rep)


		except:
		 	if '<representations total_record_count="0"/>' in my_api.xml_response_data:
		 	 	num_rep = 0


		if title in title_dict:
			title_dict[title][mms] = {"ie":ies, "holdings":holdings, "series":series}
			 
		else:
		 	mms_data = {"ie":ies, "holdings":holdings, "series":series}
		 	title_data= {mms:mms_data}
		 	title_dict[title] = title_data
		#print(title_dict)
		with open(json_file_path, 'w') as f:
			json.dump(title_dict, f)
def make_json_from_text(textfilename):
	my_text_dict = {}
	with open(textfilename,"r") as f:
		data = f.read()
	for el in data.split("\n")[:-1]:
		my_text_dict[el.split("{")[0]]="{"+"{".join(el.split("{")[1:])

	newfilename = textfilename.replace(".txt",".json")
	with open(newfilename, 'w') as f:
		json.dump(my_text_dict, f)
	return newfilename

def check_for_no_items_rep():
	"""This function is checking for incomplete records"""
	print("#"*50)
	print("Checking for incomplete items or representations")
	print("#"*50)

	with open(json_file_path, "r") as f:
		title_dict = json.load(f)
	count_empty_holdings=0
	count_no_items =0
	count_no_rep=0
	no_holdings = []
	no_items = []
	no_reps = []
	no_holdings_but_rep={}
	no_hold_rep_count = 0
	no_holdings_no_reps = {}
	count_no_hold_no_reps =0
	no_rep_but_holding = {}
	count_no_rep_hold = 0 
	for title in  title_dict.keys():
		for mms in title_dict[title].keys():
			flag_hold= False
			flag_rep = False
			if len(title_dict[title][mms]["holdings"])==0:
				#print(title)
				# count_empty_holdings+=1
				# no_holdings.append(title_dict[title])
				flag_hold=True
				# print(title_dict[title][mms])
			else:

				for hold in title_dict[title][mms]["holdings"]:
					if len(title_dict[title][mms]["holdings"][hold])==0:
						# print(title)
						# print(title_dict[title][mms])

						count_no_items+=1
						no_items.append(title_dict[title])
			#print(title_dict[title][mms]["ie"])
			if len(title_dict[title][mms]["ie"])==0:
				flag_rep = True
				# print(title)
				# print(title_dict[title][mms])
				# count_no_rep +=1
				# no_reps.append(title_dict[title])
		
			if flag_hold and not flag_rep:
					no_holdings_but_rep[title]=title_dict[title]
					no_hold_rep_count+=1
			if flag_hold and flag_rep:
					no_holdings_no_reps[title]=title_dict[title]
					count_no_hold_no_reps+=1
			if not flag_hold and flag_rep:
					no_rep_but_holding[title] = title_dict[title]
					count_no_rep_hold+=1
	print("!"*50)
	print("No items: ",count_no_items)
	print("!"*50)
	if count_no_items>0:
		for el in no_items:
			print(no_items[el])
	print("!"*50)
	print("No holdings but representation exists: ", no_hold_rep_count)
	print("!"*50)
	if no_hold_rep_count>0:
		for el in no_holdings_but_rep:
			print(no_holdings_but_rep[el])
			with open(no_hold_rep_file,"a",encoding="UTF-8") as f:
				f.write(el+"\t"+repr(no_holdings_but_rep[el])+"\n")
	
	print("!"*50)
	print("No representation and no holding: ",count_no_hold_no_reps)
	print("!"*50)
	if count_no_hold_no_reps>0:
		for el in no_holdings_no_reps:
			print(no_holdings_no_reps[el])
			with open(no_rep_no_hold_file,"a",encoding="UTF-8") as f:
				f.write(el+"\t"+repr(no_holdings_no_reps[el])+"\n")

	print("!"*50)
	print("Holding exists and not representation: ",count_no_rep_hold)
	print("!"*50)
	if count_no_rep_hold>0:
		for el in no_rep_but_holding:
			print(no_rep_but_holding[el])
			with open(no_rep_hold_file ,"a",encoding="UTF-8") as f:
				f.write(el+"\t"+repr(no_rep_but_holding[el])+"\n")

def prepare_ies_for_delete():
	my_dict = {
'9918809571102836': {'ie': {'IE44626009': 'Papercuts episode: Three Women'}, 'holdings': {'22330953710002836': ['23330953690002836']}, 'series': 'Papercuts; July 25, 2019.'}, 
'9918845673402836': {'ie': {'IE48445793': 'Episode 52 - Nicole Gaffney (Carb On Carb, Girls Rock)'}, 'holdings': {'22335025350002836': ['23335025330002836']}, 'series': "Musician's map (Series); episode 52."},
'9919054567602836': {'ie': {'IE65247625': 'At Our Most F***ed (w/ Maddy &amp; Simone)'}, 'holdings': {'22361183000002836': ['23361193000002836']}, 'series': 'Every stupid question ; June 07, 2019.'},
'9919067834302836': {'ie': {'IE67699974': 'The Angus Dunn Podcast Episode 3: Chris Starr, Yoga Teacher'}, 'holdings': {'22362571400002836': ['23362561390002836']}, 'series': 'Angus Dunn podcast ; episode 3.'},
'9919077072802836': {'ie': {'IE68769889': 'The Laws Of Robotics'}, 'holdings': {'22363708950002836': ['23363719430002836']}, 'series': 'Human-robot interaction podcast ; February 18, 2021.'},
'9919235169902836': {'ie': {'IE83043283': 'Shift work'}, 'holdings': {'22385627480002836': ['23385627460002836']}, 'series': 'Goodfellow podcast ; October 06, 2021.'},
'9919235170002836': {'ie': {'IE83043271': 'Libido in the menopausal woman'}, 'holdings': {'22385636910002836': ['23385636890002836']}, 'series': 'Goodfellow podcast ; October 19, 2021.'},
'9919235170102836': {'ie': {'IE83043213': 'Assisted dying'}, 'holdings': {'22385636960002836': ['23385636940002836']}, 'series': 'Goodfellow podcast ; December 01, 2021.'}, 
'9919235259302836': {'ie': {'IE83043345': 'Managing depressed and anxious patients'}, 'holdings': {'22385636830002836': ['23385636810002836']}, 'series': 'Goodfellow podcast ; August 11, 2021.'},
'9919235258902836': {'ie': {'IE83043378': 'Reversing type two diabetes'}, 'holdings': {'22385636760002836': ['23385627320002836']}, 'series': 'Goodfellow podcast ; June 16, 2021.'},
'9919235258802836': {'ie': {'IE83043387': 'Lifestyle medicine'}, 'holdings': {'22385636740002836': ['23385636720002836']}, 'series': 'Goodfellow podcast ; June 02, 2021.'},
'9919235259002836': {'ie': {'IE83043369': 'Clozapine and primary care'}, 'holdings': {'22385627360002836': ['23385627330002836']}, 'series': 'Goodfellow podcast ; June 30, 2021.'}, 
'9919235259102836': {'ie': {'IE83043357': "Endometriosis - what's new and important"}, 'holdings': {'22385636790002836': ['23385636770002836']}, 'series': 'Goodfellow podcast ; July 13, 2021.'},  
'9919235259202836': {'ie': {'IE83043348': "Alzheimer's Disease: Metabolic strategies to slow and reverse progression"}, 'holdings': {'22385627380002836': ['23385636800002836']}, 'series': 'Goodfellow podcast ; July 28, 2021.'}, 
'9919235259402836': {'ie': {'IE83043336': 'Covid vaccine hesitancy'}, 'holdings': {'22385627410002836': ['23385627390002836']}, 'series': 'Goodfellow podcast ; August 18, 2021.'}, 
'9919235259502836': {'ie': {'IE83043304': 'Teenage brain development'}, 'holdings': {'22385636860002836': ['23385636840002836']}, 'series': 'Goodfellow podcast ; August 25, 2021.'},
'9919235259602836': {'ie': {'IE83043298': 'A guide on using Dulaglutide for T2D'}, 'holdings': {'22385627440002836': ['23385627420002836']}, 'series': 'Goodfellow podcast ; September 08, 2021.'}, 
'9919235259702836': {'ie': {'IE83043292': 'Diabetes management: a case-based discussion'}, 'holdings': {'22385636880002836': ['23385627450002836']}, 'series': 'Goodfellow podcast ; September 22, 2021.'}, 
'9919235259802836': {'ie': {'IE83043255': 'Last days of life'}, 'holdings': {'22385636930002836': ['23385627490002836']}, 'series': 'Goodfellow podcast ; November 03, 2021.'}, 
'9919235259902836': {'ie': {'IE83043226': 'Ventilation and filtration of clinics during COVID'}, 'holdings': {'22385627520002836': ['23385627500002836']}, 'series': 'Goodfellow podcast ; November 16, 2021.'}, 
'9919235260002836': {'ie': {'IE83043204': 'Mild cognitive impairment'}, 'holdings': {'22385636990002836': ['23385636970002836']}, 'series': 'Goodfellow podcast ; December 15, 2021.'}, 
'9919235257802836': {'ie': {'IE83042161': 'The Elephant Hates You'}, 'holdings': {'22385565780002836': ['23385575960002836']}, 'series': 'Chris and Sam podcast ; June 18, 2022.'},
'9919235169302836': {'ie': {'IE83042158': 'Learning to Walk Again'}, 'holdings': {'22385575980002836': ['23385565790002836']}, 'series': 'Chris and Sam podcast ; June 22, 2022.'},
'9919235257302836': {'ie': {'IE83042582': 'Carrion Company S2-8 Return of the Vulture'}, 'holdings': {'22385630920002836': ['23385630900002836']}, 'series': 'Mud & blood ; June 20, 2022.'},
'9919235169202836': {'ie': {'IE83042164': 'Nighttime Solar Panel'}, 'holdings': {'22385565760002836': ['23385565740002836']}, 'series': 'Chris and Sam podcast ; June 11, 2022.'},
'9919240122002836': {'ie': {'IE83328341': 'The bedroom murders: love, jealousy and deadly revenge in South Auckland'}, 'holdings': {'22385942020002836': ['23385942000002836']}, 'series': 'Moment in crime ; August 05, 2022.'}}

	for mms in my_dict:

		print(list(my_dict[mms]["ie"].keys())[0]+"\t'"+mms+"\t"+list(my_dict[mms]["ie"].items())[0][1])

def check_for_double_item_rep():
	"Checks for duplicates in items, holdings and representations"
	print("#"*50)
	print("Checking for duplicates items and representations within each mms")
	print("#"*50)
	with open(json_file_path, "r") as f:
		title_dict = json.load(f)
		count_double_holdings=0
		count_double_items=0
		count_double_ies = 0
		double_holdings = []
		double_items = []
		double_ies = []
	for title in  title_dict.keys():
		for mms in title_dict[title].keys():
			# print(title_dict[title][mms]["holdings"])
			# print(len(title_dict[title][mms]["holdings"]))
			if len(title_dict[title][mms]["holdings"])>1:
				count_double_holdings+=1
				double_holdings.append(title_dict[title][mms]["holdings"])
			for hold in title_dict[title][mms]["holdings"]:
				# print((title_dict[title][mms]["holdings"][hold]))
				# print(len(title_dict[title][mms]["holdings"][hold]))
				if len(title_dict[title][mms]["holdings"][hold])>1:
					count_double_items+=1
					double_items.append(title_dict[title][mms]["holdings"][hold])
			if len(title_dict[title][mms]["ie"])>1:
				count_double_ies +=1
				# print(title_dict[title][mms]["series"])
				# print(title_dict[title][mms]["ie"])
				double_ies.append(title_dict[title][mms]["ie"])
	print("Duplicated holdings: ",count_double_holdings)
	if count_double_holdings>0:
		print(double_holdings)
	print("Duplicated items: ",count_double_items)
	if count_double_items>0:
			print(double_items)
	print("Duplicated ies: ",count_double_ies)
	if count_double_ies>0:
		print(double_ies)


def clean_dictionary_from_good_records():
	print("Cleaning dictionary from good item")


	with open(json_file_path, "r") as f:
		title_dict = json.load(f)
		count_good_records = 0
	count_all = len(title_dict)
	new_dict = dict(title_dict)
	for title in  title_dict.keys():
		if len(title_dict[title])==1:
			#print(title_dict[title])
			new_dict.pop(title)
			count_good_records +=1

		else:
			flag_series= False
			list_series = []
			for mms in title_dict[title]:
				if title_dict[title][mms]["series"] not in list_series:
					list_series.append(title_dict[title][mms]["series"])
				else:
					flag_series = True
			if not flag_series:
				new_dict.pop(title)
				count_good_records +=1
				print("Good record")
				print(title)
				print(title_dict[title])
	print("All records - ",count_all)
	print("Good_records - ",count_good_records)
	print("To work on - ", count_all - count_good_records)
	print(json_second_path)
	with open(json_second_path,"w") as f:
		json.dump(new_dict,f)
	for el in new_dict.keys():
		print(el)
		print(new_dict[el])

def delete_mms_with_one_good_and_others_empty():
		
		count_to_delete=0
		with open(json_second_path,"r") as f:
			title_dict = json.load(f)
		new_dict = dict(title_dict)
		for title in  title_dict.keys():
			#print(title_dict[title])
			list_full = []
			list_empty = []
			
			for mms in title_dict[title]:
				# print(title_dict[title][mms]["ie"].items())
				# print(len(title_dict[title][mms]["ie"].items()))
				# print(title_dict[title][mms]["ie"])
				# print(len(title_dict[title][mms]["ie"]))

				if len(title_dict[title][mms]["ie"].items())==0:

					# print("Empty")
					# print(title)
					# print(title_dict[title][mms])
					list_empty.append(mms)
				else:
					# print("Full")
					# print(title)
					# print(title_dict[title][mms])
					list_full.append(mms)
			if len(list_full) == 1:
				print(list_empty)
				for mms in list_empty:
					my_api.delete_bib(mms)
					if str(my_api.status_code).startswith("2"):
						print(mms, " - deleted")

						try:
							new_dict.pop(title)
						except:
							pass
					else:
						# print(my_api.xml_response_data)
						my_api.get_holdings(mms)
						#print(my_api.xml_response_data)
						holding_ids = re.findall(r"<holding_id>(.*?)</holding_id>", my_api.xml_response_data)
						#print(holding_ids)

						for holding_id in holding_ids:
							item_count=0
							my_api.get_items(mms,holding_id,{"limit":"100"})
							item_count = re.findall(r'_count="(.*?)">',my_api.xml_response_data)[0]
							for i in range((int(item_count)//100)+2):
								my_api.get_items(mms,holding_id,{"limit":"100","offset":100*i})
								items = re.findall(r"<pid>(.*?)</pid>",my_api.xml_response_data)
								print(items)
								for item in items:
									my_api.get_item(mms,holding_id, item)
									new_item_data = my_api.xml_response_data.replace('<committed_to_retain desc="Yes">true</committed_to_retain>','<committed_to_retain desc="No">false</committed_to_retain>')
									po_line = re.findall(r"<po_line>(.*?)</po_line>", my_api.xml_response_data)[0]
									new_item_data = new_item_data.replace("<po_line>"+po_line+"</po_line>","")
									try:
										my_api.delete_po_line(po_line)
						#				print(my_api.xml_response_data)
									except:
										print(my_api.xml_response_data)


									my_api.update_item(mms,  holding_id, item, new_item_data )
						#			print(my_api.xml_response_data)
									if str(my_api.status_code).startswith("2"):
										my_api.delete_item(mms,  holding_id, item)
						#				print(my_api.xml_response_data)
										if str(my_api.status_code).startswith("2"):
											print(mms, holding_id, item, "- item deleted")
							my_api.delete_holding(mms, holding_id)
							if str(my_api.status_code).startswith("2"):
								print(mms, holding_id, "- holding deleted")
					my_api.delete_bib(mms)
					if str(my_api.status_code).startswith("2"):
						print(mms, "- bib deleted")
						try:
							new_dict.pop(title)
						except:
							pass
					else:
						print(my_api.xml_response_data)

		print(len(new_dict))
		for el in new_dict:
			print(el)
			print(new_dict[el])
		with open(json_third_path,"w") as f:
			json.dump(new_dict, f)

# def make_csv_for_deleting():
# 	with open (json_third_path,"r") as f:
# 		title_dict = json.load(f)
# 	for title in  title_dict.keys():
# 		for mms in title_dict[title]:
# 			print(mms)


def get_records_from_set(set_id):

	my_api.get_set_members(set_id,{"limit":"100"})
	print(my_api.xml_response_data)
	total_count = re.findall(r'<members total_record_count="(.*?)">',my_api.xml_response_data)[0]
	for i in range((int(total_count)//100)+2):
			my_api.get_set_members(set_id,{"limit":"100","offset":100*i})
			mms_ids = re.findall(r"<id>(.*?)</id>",my_api.xml_response_data)
			for mms in mms_ids:
				try:
					my_api.get_bib(mms)
				except:
					my_api.get_bib(mms)
				print(my_api.xml_response_data)
				my_data = "".join(my_api.xml_response_data.split("\n"))
				if "Record automatically derived from RSS feed" in my_api.xml_response_data and not "The Spinoff presents SUPERPOD" in my_api.xml_response_data:
					with open(file_name,"a", encoding = "UTF-8") as f:
						f.write(my_data)
						f.write("\n")


def delete_if_one_good_and_others_empty(my_duplicates_entire_info):

	rep_flag = 0
	hold_flag = 0
	for title in my_duplicates_entire_info:
		rep_flag = False
		hold_flag = False
		count = 0
		mms_to_save = []
		print("="*100)
		print(title)
		for mms in my_duplicates_entire_info[title][0]:
			if my_duplicates_entire_info[title][0][mms]["hold"] == '1':
				rep_flag = True
				count +=1
				if my_duplicates_entire_info[title][0][mms]["repres"] == '1':
					hold_flag =True
					mms_to_save.append(mms)
			if count>1:
				mms_to_save.append(mms)
		if rep_flag and hold_flag:
			for mms in my_duplicates_entire_info[title][0]:
				if mms not in  mms_to_save:
					my_api.delete_bib(mms)
					print("MMS to delete:",mms)
					if str(my_api.status_code).startswith("2"):
						print("deleted")
					else:
						print(my_api.xml_response_data)
						print(my_api.status_code)

				else:
					print("MMS to keep:", mms)
		elif rep_flag  and not hold_flag:
			print("#"*100)
			print("Please check if physical items were made for this title")
			print(title)
			print(my_duplicates_entire_info[title][0])
			print("#"*100)
		elif not rep_flag and hold_flag:
			print("#"*100)
			print("Please check if the title has a representation")
			print(title)
			print(my_duplicates_entire_info[title][0])
			print("#"*100)
		elif not rep_flag and not hold_flag:
			print("#"*100)
			print("Please check if the title has representation and item")
			print(title)
			print(my_duplicates_entire_info[title][0])
			print("#"*100)

		if len(mms_to_save) >1:
			print("#"*100)
			print("These titles were not deleted due to discovering of multiple representations for multiple records. Please check manually if they duplicates")
			print(title)
			print(my_duplicates_entire_info[title][0])
			print("#"*100)



def get_mark_from_file():

	my_duplicates = {}
	with open(file_name,"r", encoding = "UTF-8") as f:
		data = f.read()
	count = 0
	my_list = []
	for el in data.split("\n")[:-1]:
		if el.startswith("<?xml "):
			my_list.append(el)
			count+=1
	# print(len(my_list))
	my_new_list = list(set(my_list))
	# print(len(my_new_list))
	for el in my_new_list:
		try:
			record = parse_xml_to_array(io.StringIO(el))[0]
			# print(record)
		except:
			el = el + "</subfield></datafield></record></bib>"
			record = parse_xml_to_array(io.StringIO(el))[0]

		# print(record)
		mms_id = record["001"].value()
		title = record["245"]["a"]
		if title in my_duplicates.keys():
			my_duplicates[title] +=[mms_id]
		else:
			my_duplicates[title] = [mms_id]

	my_duplicates_entire_info={}
	for el in my_duplicates:
		small_dict = {}
		if len(my_duplicates[el])>2:
			if not "The Spinoff presents SUPERPOD" in el:
				print("#"*50)
				print(el)
				for mms in my_duplicates[el]:
					print("-"*50)
					print(mms)
					my_api.get_representations(mms,{"limit":"100"})
					try:
						rep_count = re.findall(r'count="(.*?)">',my_api.xml_response_data)[0]
					except:
						rep_count = 0
					print("Rep count: ", rep_count)
					my_api.get_holdings(mms, {"limit":"100"})
					try:
						hold_count = re.findall(r'count="(.*?)">',my_api.xml_response_data)[0]
					except:
						hold_count = 0
					print("Hold count: ",  hold_count)
					small_dict[mms] = {"hold":hold_count, "repres":rep_count}
					if el in my_duplicates_entire_info.keys():
						my_duplicates_entire_info[el]+=[small_dict]
					else:
						my_duplicates_entire_info[el]=[small_dict]
	return my_duplicates_entire_info


def main():
	#Old   functions###
	#set_id = "11320400410002836"
	#get_records_from_set(set_id)

	#Main script
###########################################################
	#!!!pick up one of the functions!!

	#--working with Alma apis and set--
	#target_set(cleaning=False)

	#--working with spreadsheet--------
	#parse_spreadsheet(cleaning = False)
############################################################
	#check_for_double_item_rep()
	check_for_no_items_rep()
	# noreholjson = make_json_from_text(no_rep_hold_file)
	# noholrepjson = make_json_from_text(no_hold_rep_file)
	# noholdnorepsjson= make_json_from_text(no_rep_no_hold_file)
	
	#prepare_ies_for_delete()
	#clean_dictionary_from_good_records()
	#delete_mms_with_one_good_and_others_empty()


	##TO DO
	#delete bib from forth (manually created by librarian from (mostly from no reps reports)) file (text)


	
if __name__ == '__main__':
	main()