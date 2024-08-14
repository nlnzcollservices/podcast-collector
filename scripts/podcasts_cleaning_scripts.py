import sys
import re
import io
import os
import csv
import json
from pathlib import Path
from datetime import datetime as dt
from pymarc import parse_xml_to_array
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
import openpyxl
print(os.getcwd())
from settings import report_folder, assets_folder
import feedparser
import dateparser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
import openpyxl
import random
from fuzzywuzzy import fuzz
from podcast_dict import podcasts_dict
sys.path.insert(0, r"Y:\ndha\pre-deposit_prod\LD_working\alma_tools")
from alma_tools import AlmaTools
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
base_dir = os.path.abspath(os.path.join(parent_dir, os.pardir))
sys.path.insert(0, parent_dir)
time_string = "_"+dt.now().strftime("%m_%Y")
cleaning_report_folder = "cleaning_report"+time_string



class PodcastCleaningPipeline:

    def __init__(self, report_folder=os.getcwd()):

        self.report_folder = report_folder

        
        print(report_folder)


        time_string = "_" + dt.now().strftime("%m_%Y")
        self.cleaning_report_folder = "cleaning_report" + time_string
        print(report_folder)
        self.cleaning_report_folder_path = os.path.join(report_folder, self.cleaning_report_folder)

        if not os.path.isdir(self.cleaning_report_folder_path):
            os.mkdir(self.cleaning_report_folder_path)

        self.json_file_path = os.path.join(self.cleaning_report_folder_path, "podcast_main_cleaning_report.json")
        self.json_second_path = os.path.join(self.cleaning_report_folder_path, "podcast_possible_dups.json")
        self.json_third_path = os.path.join(self.cleaning_report_folder_path, "podcasts_most_problematic.json")
        self.no_rep_no_hold_file = os.path.join(self.cleaning_report_folder_path, "no_reps_no_holdings.txt")
        self.no_rep_hold_file = os.path.join(self.cleaning_report_folder_path, "no_reps_holdings.txt")
        self.no_hold_rep_file = os.path.join(self.cleaning_report_folder_path, "no_hold_rep_file.txt")
        self.report_cleaning_temp = os.path.join(self.cleaning_report_folder_path, "report_cleaning_temp_count.txt")
        self.rss_sprsheet_path = os.path.join(self.cleaning_report_folder_path, "rss_check_results.xlsx")


    def cleanup(self):

        """This function deletes json and temporary numbering file"""

        try:
            os.remove(self.json_file_path)
            os.remove(self.report_cleaning_temp)
        except Exception as e:
            print(str(e))

   
   
    def clean_dictionary_from_good_records(self, json_file_path):

        """The function removes correct records from dictionary and saves incorrect to "json_possible_dups.json"

        Parameters:
            json_file_path(str) - path to the main json file
        Returns:
            results (str) - function results for printing in the GUI log window
         """

        print("Remove correct results from dictionary")


        with open(json_file_path, "r") as f:
            title_dict = json.load(f)
            count_good_records = 0
        count_all = len(title_dict)
        new_dict = dict(title_dict)
        for title in  title_dict.keys():
            if len(title_dict[title])==1:
                #print(title_dict[title])
                new_dict.pop(title)
                count_good_records +=1

            else:
                flag_series= False
                list_series = []
                for mms in title_dict[title]:
                    if title_dict[title][mms]["series"] not in list_series:
                        list_series.append(title_dict[title][mms]["series"])
                    else:
                        flag_series = True
                if not flag_series:
                    new_dict.pop(title)
                    count_good_records +=1
                    print("Good record")
                    print(title)
                    print(title_dict[title])

        results = ""
        print("All records - ",count_all)
        results += "All records - "
        results +=str(count_all)
        results += "\n"
        print("Good_records - ",count_good_records)
        results += "Good_records - "
        results +=str(count_good_records)
        results += "\n"
        print("To work on - ", count_all - count_good_records)
        results += "To work on - "
        results +=str(count_all - count_good_records)
        results += "\n"

        print(self.json_second_path)
        with open(self.json_second_path,"w") as f:
            json.dump(new_dict,f)
        return results

    def delete_bib(self, mms , my_api):

        """This function deletes bibs
        Parameters:
            mms (str) - Alma mms id
            my_api(obj) - Alma tools object

        """

        my_api.get_bib(mms)
        try:
            my_rec  = parse_xml_to_array(io.StringIO(my_api.xml_response_data))[0]
            my_rec.remove_fields("042")
            my_rec= record_to_xml(my_rec)
            start_xml = r'<?xml version="1.0" encoding="UTF-8"?><bib><record_format>marc21</record_format><suppress_from_publishing>false</suppress_from_publishing><suppress_from_external_search>false</suppress_from_external_search><sync_with_oclc>BIBS</sync_with_oclc>'
            end_xml = '</bib>'
            my_rec = str(my_rec).replace("\\n", "\n").replace("\\", "")
            my_xml =  start_xml + my_rec +end_xml
            my_api.update_bib(mms, my_xml)
            my_api.delete_bib(mms)
            print(mms, "deleted")
        except:
            print(mms, "could not delete")
    
    def delete_mms_with_one_good_and_others_empty(self, json_second_path):

        """This function deletes empty bib records, if full duplicate exists

        Parameters:
            json_file_path(str) - path to the main json file
        Returns:
            results (str) - function results for printing in the GUI log window
         """
        
        my_api = AlmaTools("prod")

        count_to_delete=0
        with open(self.json_second_path,"r") as f:
            title_dict = json.load(f)
        new_dict = dict(title_dict)
        for title in  title_dict.keys():
            #print(title_dict[title])
            list_full = []
            list_empty = []
            
            for mms in title_dict[title]:
                if len(title_dict[title][mms]["ie"].items())==0:
                    list_empty.append(mms)
                else:
                    list_full.append(mms)
            if len(list_full) == 1:
                print(list_empty)
                for mms in list_empty:
                    self.delete_bib(mms , my_api)
                    print(my_api.xml_response_data)
                    if str(my_api.status_code).startswith("2"):
                        print(mms, " - deleted")
                        try:
                            new_dict.pop(title)
                        except:
                            pass
                    else:
                        print(my_api.xml_response_data)
                        my_api.get_holdings(mms)
                        holding_ids = re.findall(r"<holding_id>(.*?)</holding_id>", my_api.xml_response_data)

                        for holding_id in holding_ids:
                            item_count=0
                            my_api.get_items(mms,holding_id,{"limit":"100"})
                            try:
                                item_count = re.findall(r'_count="(.*?)">',my_api.xml_response_data)[0]
                            except:
                                item_count = re.findall(r'_count="(.*?)"/>',my_api.xml_response_data)[0]
                            for i in range((int(item_count)//100)+2):
                                my_api.get_items(mms,holding_id,{"limit":"100","offset":100*i})
                                items = re.findall(r"<pid>(.*?)</pid>",my_api.xml_response_data)
                                print(items)
                                for item in items:
                                    my_api.get_item(mms,holding_id, item)
                                    new_item_data = my_api.xml_response_data.replace('<committed_to_retain desc="Yes">true</committed_to_retain>','<committed_to_retain desc="No">false</committed_to_retain>')
                                    po_line = re.findall(r"<po_line>(.*?)</po_line>", my_api.xml_response_data)[0]
                                    new_item_data = new_item_data.replace("<po_line>"+po_line+"</po_line>","")
                                    try:
                                        my_api.delete_po_line(po_line)
                                    except:
                                        print(my_api.xml_response_data)


                                    my_api.update_item(mms,  holding_id, item, new_item_data )
                                    if str(my_api.status_code).startswith("2"):
                                        my_api.delete_item(mms,  holding_id, item)
                                        if str(my_api.status_code).startswith("2"):
                                            print(mms, holding_id, item, "- item deleted")
                            my_api.delete_holding(mms, holding_id)
                            if str(my_api.status_code).startswith("2"):
                                print(mms, holding_id, "- holding deleted")

                    if str(my_api.status_code).startswith("2"):
                        print(mms, "- bib deleted")
                        try:
                            new_dict.pop(title)
                        except:
                            pass
                    else:
                        print(my_api.xml_response_data)

        print("Number of problematic: ",len(new_dict))
        results = "Number of problematic bibs: " + str(len(new_dict)) + "\n"
        for el in new_dict:
            print(el)
            results += el + "\n" + repr(new_dict[el]) +"\n"
            print(new_dict[el])
        with open(self.json_third_path,"w") as f:
            json.dump(new_dict, f)
        return results



    def check_for_double_item_rep(self ,json_file_path):


        print("#"*50)
        print("Checking for duplicates items and representations within each mms")
        print("#"*50)
        with open(json_file_path, "r") as f:
            title_dict = json.load(f)
            count_double_holdings=0
            count_double_items=0
            count_double_ies = 0
            double_holdings = []
            double_items = []
            double_ies = []
        for title in  title_dict.keys():
            for mms in title_dict[title].keys():
                # print(title_dict[title][mms]["holdings"])
                # print(len(title_dict[title][mms]["holdings"]))
                if len(title_dict[title][mms]["holdings"])>1:
                    count_double_holdings+=1
                    double_holdings.append(title_dict[title][mms]["holdings"])
                for hold in title_dict[title][mms]["holdings"]:
                    # print((title_dict[title][mms]["holdings"][hold]))
                    # print(len(title_dict[title][mms]["holdings"][hold]))
                    if len(title_dict[title][mms]["holdings"][hold])>1:
                        count_double_items+=1
                        double_items.append(title_dict[title][mms]["holdings"][hold])
                if len(title_dict[title][mms]["ie"])>1:
                    count_double_ies +=1
                    # print(title_dict[title][mms]["series"])
                    # print(title_dict[title][mms]["ie"])
                    double_ies.append(title_dict[title][mms]["ie"])
        results = ""
        print("Duplicated holdings: ",count_double_holdings)
        results+="Duplicated holdings: "
        results+=str(count_double_holdings)
        results+="\n"
        print("Duplicated items: ",count_double_items)
        results+="Duplicated items: "
        results+=str(count_double_items)
        results+="\n"
        print("Duplicated ies: ",count_double_ies)
        results+="Duplicated reps: "
        results+=str(count_double_ies)
        results+="\n"
        return results

    def make_json_from_text(self, textfilename):

        my_text_dict = {}
        with open(textfilename,"r", encoding = "UTF-8") as f:
            data = f.read()
        for el in data.split("\n")[:-1]:
            my_text_dict[el.split("{")[0]]="{"+"{".join(el.split("{")[1:])

        newfilename = textfilename.replace(".txt",".json")
        with open(newfilename, 'w') as f:
            json.dump(my_text_dict, f)
        return newfilename
       
    def check_for_no_items_rep(self, json_file_path):

        """This function is checking for incomplete records"""
        print("#"*50)
        print("Checking for incomplete items or representations")
        print("#"*50)

        with open(json_file_path, "r") as f:
            title_dict = json.load(f)
        count_empty_holdings=0
        count_no_items =0
        count_no_rep=0
        no_holdings = []
        no_items = []
        no_reps = []
        no_holdings_but_rep={}
        no_hold_rep_count = 0
        no_holdings_no_reps = {}
        count_no_hold_no_reps =0
        no_rep_but_holding = {}
        count_no_rep_hold = 0 
        for title in  title_dict.keys():
            for mms in title_dict[title].keys():
                flag_hold= False
                flag_rep = False
                if len(title_dict[title][mms]["holdings"])==0:
                    flag_hold=True
                else:

                    for hold in title_dict[title][mms]["holdings"]:
                        if len(title_dict[title][mms]["holdings"][hold])==0:
                            count_no_items+=1
                            no_items.append(title_dict[title])
                if len(title_dict[title][mms]["ie"])==0:
                    flag_rep = True

                if flag_hold and not flag_rep:
                        no_holdings_but_rep[title]=title_dict[title]
                        no_hold_rep_count+=1
                if flag_hold and flag_rep:
                        no_holdings_no_reps[title]=title_dict[title]
                        count_no_hold_no_reps+=1
                if not flag_hold and flag_rep:
                        no_rep_but_holding[title] = title_dict[title]
                        count_no_rep_hold+=1
        results = ""
        print("!"*50)
        print("No items: ",count_no_items)
        print("!"*50)
        if count_no_items>0:
            for el in no_items:
                print(no_items[el])
        results +="No items: "
        results+= str(count_no_items)
        results+="\n"
        print("!"*50)
        print("No holdings but representation exists: ", no_hold_rep_count)
        print("!"*50)
        if no_hold_rep_count>0:
            for el in no_holdings_but_rep:
                print(no_holdings_but_rep[el])
                with open(self.no_hold_rep_file,"a",encoding="UTF-8") as f:
                    f.write(el+"\t"+repr(no_holdings_but_rep[el])+"\n")
        results +="No holdings but representation exists: "
        results+= str(no_hold_rep_count)
        results+="\n"
        try:
            noreholjson = self.make_json_from_text(self.no_hold_rep_file)
        except:
            pass
        print("!"*50)
        print("No representation and no holding: ",count_no_hold_no_reps)
        print("!"*50)
        if count_no_hold_no_reps>0:
            for el in no_holdings_no_reps:
                print(no_holdings_no_reps[el])
                with open(self.no_rep_no_hold_file,"a",encoding="UTF-8") as f:
                    f.write(el+"\t"+repr(no_holdings_no_reps[el])+"\n")
        results +="No representation and no holding: "
        results+= str(count_no_hold_no_reps)
        results+="\n"
        try:
            noholrepjson = self.make_json_from_text(self.no_rep_no_hold_file)
        except:
            pass

        print("!"*50)
        print("Holding exists but no representation: ",count_no_rep_hold)
        print("!"*50)
        if count_no_rep_hold>0:
            for el in no_rep_but_holding:
                print(no_rep_but_holding[el])
                with open(self.no_rep_hold_file ,"a",encoding="UTF-8") as f:
                    f.write(el+"\t"+repr(self.no_rep_but_holding[el])+"\n")
        results +="Holding exists and no representation: "
        results+= str(count_no_rep_hold)
        results+="\n"
        try:
            noholdnorepsjson= self.make_json_from_text(self.no_rep_hold_file)
        except:
            pass
        return results


    def load_title_dict(self):

        try:
            with open(self.json_file_path, 'r') as f:
                title_dict = json.load(f)
        except Exception as e:
            print(str(e))
            title_dict = {}
        return title_dict

    def save_title_dict(self, title_dict):

        with open(self.json_file_path, 'w') as f:
            json.dump(title_dict, f)

    def load_my_number(self):
        try:
            with open(self.report_cleaning_temp, "r") as f:
                my_number = int(f.read())
        except:
            my_number = 0
        return my_number

    def save_my_number(self, my_number):
        with open(self.report_cleaning_temp, "w") as f:
            f.write(str(my_number))



    def prepare_ies_for_delete(self,fl, output_csv_folder):
        """
        Reads a JSON file, processes its content, and writes specific data to a CSV file.

        Parameters:
        - fl (str): Path to the input JSON file.
        - output_csv_path (str): Path where the output CSV file will be saved.
        """
        print(f"Processing file: {fl}")

        # Ensure the output directory exists
        output_csv_path = os.path.join(output_csv_folder,"for_ndha_form.csv")
        os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)

        with open(fl, "r") as f:
            json_dict = json.load(f)

        # Open the CSV file for writing
        with open(output_csv_path, "w", newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile, delimiter='\t')
            # Write the header row
            csv_writer.writerow(["IE Key", "MMS ID", "IE Value"])

            for el in json_dict:
                my_dict = json_dict[el]
                for mms in my_dict:
                    try:
                        ie_key = list(my_dict[mms]["ie"].keys())[0]
                        ie_value = list(my_dict[mms]["ie"].values())[0]
                        # Write the row to CSV
                        csv_writer.writerow([ie_key, mms, ie_value])
                        print(f"{ie_key}\t{mms}\t{ie_value}")
                    except Exception as e:
                        print(f"Error processing MMS ID {mms}: {e}")
                        continue

        print(f"Data has been written to {output_csv_path}")

    def parse_spreadsheet(self, sprsh_path = os.path.join(os.getcwd()), cleaning = True, mms_index="AB", serial_index="D", title_index="E"):

        """This function is parsing the spreadsheet exported from Alma , parses it and produces json main file


        Parameters:
            sprsh_path (str) - path to the spreadheet
            cleaning (bool) - remove previous results for current month, default value is True
            mms_index (str) - name of column in Excel spreadsheet  which contains mms , default is "AB"
            serial_index (str) - name of column in Excel spreadsheet which contains serial title, default is "D"
            title_index (str) - name of column in Excel spreadsheet which contains episode title, default is "E"

        """


        self.sprsh_path = sprsh_path
        self.cleaning = cleaning

        self.wb = load_workbook(self.sprsh_path)
        self.ws = self.wb["results"]
        self.my_api = AlmaTools("prod")
        # self.file_name = os.path.join(self.cleaning_report_folder_path, "podcasts_automated.txt")


        if not self.cleaning:
            title_dict = self.load_title_dict()
            my_number = self.load_my_number()

        else:
            title_dict = {}
            my_number = 0

        for i, row in enumerate(self.ws.iter_rows(min_row=my_number)):
            with open(self.report_cleaning_temp, "w") as f:
                f.write(str(i + my_number))

            rep_ids = []
            num_items = 0
            holdings = {}
            items_list = []
            title = row[title_index].value
            mms = row[mms_index].value
            series = row[serial_index].value
            reps = {}
            ies = {}

            self.my_api.get_holdings(mms, {"limit": "100"})
            try:
                #print(self.my_api.xml_response_data)
                num_hold = re.findall(r'<holdings total_record_count="(.*?)">', self.my_api.xml_response_data)[0]

                
                for _ in range(len(num_hold)):
                    hold_id = re.findall(r'holdings/(.*?)"', self.my_api.xml_response_data)[0]
                    self.my_api.get_items(mms, hold_id, {"limit": "100"})
                    items_list = []
                    try:

                        #print(self.my_api.xml_response_data)
                        num_items = re.findall(r'<items total_record_count="(.*?)">', self.my_api.xml_response_data)[0]
                        for _ in range(len(num_hold)):
                            item_pid = re.findall(r'<pid>(.*?)</pid>', self.my_api.xml_response_data)[0]
                            items_list.append(item_pid)
                        holdings[hold_id] = items_list
                    except Exception as e:
                        print(str(e))
                        if '<items total_record_count="0"/>' in self.my_api.xml_response_data:
                            num_items = 0
            except Exception as e:
                print(str(e))
                if '<holdings total_record_count="0"/>' in self.my_api.xml_response_data:
                    num_hold = 0

            self.my_api.get_representations(mms, {"limit": "100"})
            try:

                #print(self.my_api.xml_response_data)
                num_rep = re.findall(r'<representations total_record_count="(.*?)">', self.my_api.xml_response_data)[0]
                
                for _ in range(len(num_rep)):
                    rep_id = re.findall(r"<id>(.*?)</id>", self.my_api.xml_response_data)[0]
                    rep_ids.append(rep_id)
                    ie = re.findall(r"pubam:(.*?)</", self.my_api.xml_response_data)[0]
                    label = re.findall(r'<label>(.*?)</label', self.my_api.xml_response_data)[0]
                    ies[ie] = label
            except Exception as e:
                print(str(e))
                if '<representations total_record_count="0"/>' in self.my_api.xml_response_data:
                    num_rep = 0

            if title in title_dict:
                title_dict[title][mms] = {"ie": ies, "holdings": holdings, "series": series}
            else:
                mms_data = {"ie": ies, "holdings": holdings, "series": series}
                title_data = {mms: mms_data}
                title_dict[title] = title_data

            self.save_title_dict(title_dict)
    
    def clean_ep_title(self, title):
        title = title.replace("Episode","").replace("Ep","").replace("EP","").replace("podcast","").replace("The Chris and Sam","").replace("|","").replace("-","").replace("Taringa","")
        return title

    def clean_pod_title(self, title):
        title = title.replace("podcast","").replace("Podcast","")
        if ":" in title:
            title = title.split(":")[0]
        if "/"  in title:
            title = title.split("/")[0]
        return title

    def make_inverse_dict(self, json_file_path):


        """This function transforms the json dictionary for future use

        Parameters:
            json_file_path(str) - path to the main json file
        Returns:
            date_series_dict(dict) - transposed dictionary
         """

        date_series_dict =  {}
        series_list = []
        with open(json_file_path, "r") as f:
            title_dict = json.load(f)
        for title  in title_dict.keys():
            my_title = title.rstrip(".")
            for  mms in title_dict[title]:
                series_title = title_dict[title][mms]["series"].split(";")[0].rstrip(" ,")#+"|"+title +"|"+title_dict[title][mms]["series"].split(";")[-1].lstrip(" ")
                # print(series_title)
                title_date =  title_dict[title][mms]["series"].split(";")[-1].lstrip(" ").replace(",","").strip(".")
                series_list.append(series_title)
                if not "BYC" in series_title:
                    if not series_title in date_series_dict:
                        date_series_dict[series_title]={my_title :[title_date,mms]}
                    else:
                        date_series_dict[series_title][my_title ] = [title_date,mms]

        print(date_series_dict)
        return  date_series_dict

    def check_rss(self, rss_link,dates_list = None):

        """This function is for those podcasts which have dates in 490
        Parameters:
            dates_list(list) - list of dates of episodes in Alma
            rss_link(str) - rss link
        Returns:
            title_list (list) - list of rss episode titles

        """

        print("Checking rss feed for ",rss_link)

        d = feedparser.parse(rss_link)
        rss_list = []
        title_list  = []
        for i,elem in enumerate(d["entries"]):
            new_date = dateparser.parse(elem["published"]).strftime("%B %d %Y")
            rss_list.append(new_date)
            title_list.append([elem["title"],new_date])
            count_missing = 0
            
            if dates_list:
                if not new_date in dates_list:
                    print(elem["title"], new_date, " - missing")
                    count_missing+=1
        if count_missing == 0:
            print("Collected everything from rss")
        else:
            print(count, " - missing")

        return title_list


    def  make_spreadsheet(self, json_file_path):

        """This function checks all rss feeds and all records from json files and makes xlsx report

        Parameters:
            json_file_path(str) - path to the main json file
        Returns:
            results (str) - function results for printing in the GUI log window
         """

        wb = Workbook()
        results = ""

        thin_border = Border(left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin'))
        ws1 = wb["Sheet"]
        ws1.cell(column=1, row=ws1.max_row, value = "PODCASTS")
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="FFAC00")
        ws1.cell(column=3, row=ws1.max_row, value = "episodes").fill=my_fill
        ws1.cell(column=3, row=ws1.max_row).border = thin_border
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="C4B07B")
        ws1.cell(column=4, row=ws1.max_row, value = "mixed").fill=my_fill
        ws1.cell(column=4, row=ws1.max_row).border = thin_border
        my_fill= openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="95C8F0")
        ws1.cell(column=5, row=ws1.max_row, value = "date").fill=my_fill
        ws1.cell(column=5, row=ws1.max_row).border = thin_border



        date_series_dict = self.make_inverse_dict(json_file_path)
        for serial_tit in date_series_dict.keys():
            print(serial_tit)
            
            title_list = []
            not_used = []
            ser_tit = serial_tit.replace(" ",'_').replace(",","").replace("!","").replace("'","").replace("/","").replace(":","").replace("$","S")
            rss_link = None
            for d_tit in podcasts_dict.keys():
                if title_list ==[]:
                    if  fuzz.ratio(self.clean_pod_title(d_tit) ,  self.clean_pod_title(serial_tit))>80:
                        try:
                            rss_link = podcasts_dict[d_tit]["rss_filename"]
                            print(rss_link)
                            results+=d_tit + ": " +rss_link+" - ok\n"
                            title_list = self.check_rss( rss_link, None)
                        except:
                            print("no rss link")
                            results+=d_tit + ": " +rss_link+" - no rss link\n"
            for ti in title_list:
                not_used.append(ti[0])
            try:

                wb.create_sheet(ser_tit)
                ws = wb[ser_tit]
                hexadecimal = ''.join([random.choice('ABCDE') for i in range(6)])
                my_color = openpyxl.styles.colors.Color(rgb=hexadecimal)
                ws.sheet_properties.tabColor = my_color
                count_new = 0
                fuz_flag = False
                my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
                thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                thick_border = Border(left=Side(style='thick'), 
                         right=Side(style='thick'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thick'))
                ws.cell(column=1, row=ws.max_row, value=serial_tit).fill=my_fill
                link = "#Sheet!A1"
                ws.cell(column=1, row=ws.max_row).hyperlink = link
                ws.cell(column=1, row=ws.max_row).border = thick_border
                ws.cell(column=2, row=ws.max_row, value="Number/Date").fill=my_fill
                ws.cell(column=2, row=ws.max_row).border = thin_border
                ws.cell(column=3, row=ws.max_row, value="MMS id").fill=my_fill
                ws.cell(column=3, row=ws.max_row).border = thin_border
                ws.cell(column=4, row=ws.max_row, value="yyyymmdd").fill=my_fill
                ws.cell(column=4, row=ws.max_row).border = thin_border
                ws.cell(column=5, row=ws.max_row, value="Matched").fill=my_fill
                ws.cell(column=5, row=ws.max_row).border = thin_border
                ws.cell(column=6, row=ws.max_row, value="Date").fill=my_fill
                ws.cell(column=6, row=ws.max_row).border = thin_border
                ws.cell(column=7, row=ws.max_row, value="").fill=my_fill
                ws.cell(column=7, row=ws.max_row).border = thick_border
                ws.cell(column=8, row=ws.max_row, value=rss_link).fill=my_fill
                ws.cell(column=8, row=ws.max_row).border = thick_border
                ws.cell(column=9, row=ws.max_row, value="Date").fill=my_fill
                ws.cell(column=9, row=ws.max_row).border = thick_border

                
                ws.column_dimensions["A"].width = 20
                row_count =int(ws.max_row)+1
                all_count = 0
                date_count =0
                for el in date_series_dict[serial_tit].keys():
                    all_count +=1
                    ws.cell(column=1, row=row_count, value=el).border = thin_border
                    ws.cell(column=2, row=row_count, value = date_series_dict[serial_tit][el][0]).border = thin_border
                    ws.cell(column=3, row=row_count, value = date_series_dict[serial_tit][el][1]).border = thin_border
                    try:
                        # print(date_series_dict[serial_tit][el][0])
                        date_changed = dateparser.parse(date_series_dict[serial_tit][el][0]).strftime("%Y%m%d")
                        # print(date_changed)
                        if date_changed:
                            ws.cell(column=4, row=row_count, value = date_changed).border = thin_border
                            date_count +=1
                    except:
                        pass
                    row_count+=1
                    if title_list != []:
                        for r_tit in title_list:
                            if  fuzz.ratio(self.clean_ep_title(r_tit[0]) ,  self.clean_ep_title(el))>70:
                                ws.cell(column=5, row=row_count-1, value = r_tit[0]).border = thin_border
                                ws.cell(column=6, row=row_count-1, value = r_tit[1]).border = thin_border

                                for index, nu in enumerate(not_used):
                                    if nu==r_tit[0]:
                                        not_used.pop(index)


                for i,left in enumerate(not_used):
                    ws.cell(column=8, row=i+2, value = left).border = thick_border
                    for tt in title_list:
                        if tt[0]==left:
                            ws.cell(column=9, row=i+2, value = tt[1]).border = thick_border



                ws1 = wb["Sheet"]
                link = "#{}!A1".format(ser_tit)
                if date_count/all_count < 0.1:#episodes
                    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="FFAC00")
                elif date_count/all_count >=0.1 and date_count/all_count<0.9:#ep and date
                    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="C4B07B")
                elif date_count/all_count>=0.9:
                    my_fill= openpyxl.styles.fills.PatternFill(patternType='solid', fgColor="95C8F0")
                ws1.cell(column=1, row=ws1.max_row+1, value = serial_tit).fill=my_fill#.hyperlink = link
                ws1.cell(column=1, row= ws1.max_row).hyperlink = link
                ws1.column_dimensions["A"].width = 30


            except Exception as e:
                print(str(e))
                print(serial_tit)
        wb.save(self.rss_sprsheet_path)
        results += "Report saved in: " + self.rss_sprsheet_path
        return results

if __name__ == "__main__":

    json_file_path = os.path.join(r"Y:\ndha\pre-deposit_prod\LD_working\podcasts\log\reports\cleaning_report_08_2024", "podcast_main_cleaning_report.json")
    pipeline = PodcastCleaningPipeline()
    #pipeline.parse_spreadsheet()
    #pipeline.make_spreadsheet(json_file_path)
    #pipeline.check_for_double_item_rep(json_file_path)
    #pipeline.prepare_ies_for_delete(r"Y:\ndha\pre-deposit_prod\LD_working\podcasts\log\reports\cleaning_report_08_2024\podcasts_most_problematic.json", os.getcwd())