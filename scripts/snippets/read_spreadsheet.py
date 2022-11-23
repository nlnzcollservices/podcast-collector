from openpyxl import load_workbook
import os
filename = "Copy of Can_i_steal_you_podcast.xlsx"
path =r"Y:\ndha\pre-deposit_prod\LD_working\podcasts\assets"
my_string = ""
lst = []
print(type(lst))
filepath = os.path.join(path, filename)
wb = load_workbook(filepath)
ws = wb["results"]
for row in ws.iter_rows(min_row=2):
	lst.append(row[2].value)
lst = list(set(lst))
print(lst)
