import os
from pymarc import parse_xml_to_array,record_to_xml, Field, parse_xml
from settings_prod import assets_folder,  logging, template_folder,start_xml, end_xml, report_folder, pr_key#, podcast_sprsh, 
from alma_tools import AlmaTools
import io
from openpyxl import load_workbook

wb = load_workbook(os.path.join(assets_folder,"results_auto.xlsx")) 
ws = wb["results"]
my_dict ={}
for row in ws.iter_rows(min_row=2):
	podcast_name = row[3].value.split(";")[0].rstrip(" ")
	number = row[3].value.split(";")[1].lstrip(" ")
	title = row[4].value
	mms = row[26].value
	print(title)
	if title in my_dict:
		print("here2")
		my_dict[title] +=1
	else:
		print("here")
		my_dict[title] =1
count =0 
for el in my_dict:
	if my_dict[el]>1:
		count+=1
		print(el)
		print(my_dict[el])
print(count)


				
quit()
def main():

		"""Finds all mms from mms.txt file in record folder. And deletes those which have the title in title_list"""
		#title_list = ['1st birthday live show ft Andrew McDowall and Katie Wright', 'Jared Hazen', 'Gareth Thomas and the WUU2K Special', 'Rachel Grunwell', 'Hollie Woodhouse', 'Mathieu Dore', 'Josh Komen', 'Kathrine Switzer', 'Coree Woltering', 'Peter Maksimow', 'Andrew Thompson', 'Katie Wright, Riverhead Backyard Relaps', 'Chris Ord', 'Zach Miller', 'Elise Downing', 'Roger Robinson', 'Luke McCallum', 'Holly Page', 'Sam Manson', 'Gareth Morris', 'Adharanand Finn', 'Camille Herron', 'John Onate', 'Kelton Wright, Jeff Browning and Grant Guise', 'Tim Sutton', 'Weston Hill', 'Vajin Armstrong!', 'Anna Frost', 'Courtney Dauwalter', 'Dirt Church Christmas Special', 'Scott Worthington, The Revenant', 'Ruth Croft', 'Kepler Special', 'James Kuegler and Andrew McDowall LIVE', 'Marianne Elliott', 'Dylan Bowman', 'Nancy Jiang', 'Scotty Hawker', 'Lucy Bartholomew', ' Majell Backhausen', 'Crush the Cargill (Steve Tripp and Andrew Glennie)', 'Shaun Collins', 'Jeff Browning', 'Malcolm Law', 'Dr Tony Page', 'Dean Karnazes', 'Brad Dixon', 'Dawn Tuffery', 'Mel Aitken', 'Grant Guise', 'Fiona Hayvice', 'Paul Charteris', 'The Trailer']
		title_list = []
		my_alma = AlmaTools("prod")
		with open (os.path.join(report_folder, "mms.txt")) as f:
			data = f.read()
		for el in data.split("\n")[:-1]:
			#print(el)
			my_alma.get_bib(el)
			
			record =parse_xml_to_array(io.StringIO(my_alma.xml_response_data))[0]

			if record["245"]["a"].rstrip(".") in title_list:
					print("!!!!!!!!!!!!!!!!!!")
					print(record["245"]["a"].rstrip("."))
					print(record["001"].data)

				



if __name__ == '__main__':

	main()