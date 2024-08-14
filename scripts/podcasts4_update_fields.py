import os
import io
import re
import requests
import hashlib
import gspread
import dateparser
from time import mktime
from pymarc import parse_xml_to_array,record_to_xml, Field, Subfield
import time
from bs4 import BeautifulSoup
from datetime import datetime as dt
from settings import logging, file_folder, template_folder, pr_key, sb_key, logging, start_xml, end_xml
import sys
sys.path.insert(0, r"Y:\ndha\pre-deposit_prod\LD_working\alma_tools")
from alma_tools import AlmaTools
from openpyxl import load_workbook
from podcasts_database_handler import DbHandler
logger = logging.getLogger(__name__)

class Manage_fields():

	"""
	Attributes
	----------

	This class updates bibliographic record in Alma

	f347 : pymarc.Field object
		347 field 
	f500 : pymarc.Field object
		500 field
	f856 : pymarc.Field object
		856 field 
	f942 : pymarc.Field object
		942 field 
	key : str
		"prod" for production or "sb" for sandbox
	mms_id : str
		Alma mms id of bibliographic record
	mms_id_list : str
		list which contains mms ids.
	duplicate_flag : bool
		flag set False but default and become true if the field is duplicated
	update_flag: bool
		flag set False by default and becomve True if 942 added and record has to be updated

	Methods
	-------
	def __init__(self, key, mms_id_list)		
	def removing_dup_fields_add_942(self, field_num)
	def parsing_bib_xml(self)
	def cleaning_routine(self)

	"""



	def __init__(self, key,):

		self.f347 = None
		self.f500 = None
		self.f856 = None
		self.f942 = None
		self.key = key
		self.alma_key = None
		self.mms_id = None
		self.flag = False

		
			
	def removing_dup_fields_add_942(self, field_num):

		"""
		Using pymarc object for finding and removing particular duplicate field
		Parameters:
			field_num(str) - number of field

		"""
		
		fields = self.record.get_fields(field_num)
		# print(len(fields))
		# print(field_num)
		# print(self.record)
		if len(fields) != 0:
			field_dict = {}
			for ind in range(len(fields)):
				if fields[ind].value() not in field_dict:
					field_dict[fields[ind].value()] = [fields[ind]]
				else:
					field_dict[fields[ind].value()] += [fields[ind]]
					self.duplicate_flag = True

			if self.duplicate_flag:
				logger.info(field_num + "duplicated")
				self.record.remove_fields(field_num)
				for el in field_dict.keys():
					self.record.add_ordered_field(field_dict[el][0])
				logger.info("removed")


		else:
			if field_num == "942":
				if not self.short_record_flag:
					date_942 = 	(str(dt.now().strftime( '%Y-%m')))
					f942 = Field(tag = '942', indicators = ["",""], subfields = [Subfield(code='a', value="nznb {}".format(date_942))])
					self.record.add_ordered_field(f942)
					logger.info("record updated with 942")
				else:
					logger.info("the short record should not have 942")
				self.update_flag = True
						



	def parsing_bib_xml(self):

		""""Converts bib xml to pymarc. Looking for duplicates in self.field_list. Runs removing_dup_fields_add_942"""

		self.record = parse_xml_to_array(io.StringIO(self.bib_data))[0]
		self.field_list = ["347", "500", "856", "942"]
		for field_num in self.field_list:
			self.removing_dup_fields_add_942(field_num)
		print(self.record)
		self.bib_data =start_xml+str(record_to_xml(self.record)).replace("\\n", "\n").replace("\\", "")+end_xml
		logger.debug(self.bib_data)

	def cleaning_routine_old(self, mms_id_list=[]):

		"""
		Running routine for modification of bib record
		Parameters:
			mms_id_list(list) - list with mms_id record for which items already were created

		"""
		self.mms_id_list = mms_id_list
		logger.info("Updating records with 942 and removeing duplicated fields")
		sb_mms = None
		if self.key=="sb":
			sb_mms = "9918602951502836"
		self.f347 = None
		self.f500 = None
		self.f856 = None
		self.f942 = None
		self.mms_id = None
		self.bib_data = None
		for mms in self.mms_id_list:
				self.duplicate_flag = False
				self.update_flag = False
				self.mms_id = mms
				logger.info(self.mms_id)
				my_rec = AlmaTools(self.key)
				my_rec.get_bib(self.mms_id)
				self.bib_data = my_rec.xml_response_data
				self.parsing_bib_xml()
				print(self.update_flag)
				if self.update_flag:
					my_rec.update_bib(self.mms_id, self.bib_data)
					print(my_rec.status_code)
					logger.info(self.mms_id + " - updated")
					my_db =DbHandler()
					my_db.db_update_updated(self.mms_id)
				else:
					logger.info("Already has 942 field")

	def cleaning_routine(self, mms_id_list=[]):

		"""
		Running routine for modification of bib record
		Parameters:
			mms_id_list(list) - list with mms_id record for which items already were created

		"""
		#logger.setLevel("DEBUG")
		self.mms_id_list = mms_id_list
		logger.info("Updating records with 942 and removing duplicated fields")
		sb_mms = None
		if self.key=="sb":
			sb_mms = "9918602951502836"
		self.f347 = None
		self.f500 = None
		self.f856 = None
		self.f942 = None
		self.mms_id = None
		self.bib_data = None
		self.short_record_flag = False
		self.template_dict= {}
		if self.mms_id_list == []:
			# print("here")
			my_db =DbHandler()
			upd_dictionary = my_db.db_reader(["podcast_name","serial_mms","mis_mms","episode_title","holdings", "ie_num","item","updated", "serial_pol","template_name"],None, True)#episode_title", "episode_id", "date", "podcast_name","serial_pol"],None,True)
			
			for dictr in upd_dictionary:
				if 'item' in dictr.keys():
					if dictr["item"]:
						if dictr["mis_mms"]:
							self.mms_id_list.append(dictr["mis_mms"])
							if "mis_podcst_sr_" in dictr["template_name"].lower():
								self.template_dict[dictr["mis_mms"]]=True
							else:
								self.template_dict[dictr["mis_mms"]]=False
		for mms in self.mms_id_list:
			self.duplicate_flag = False
			self.update_flag = False
			self.short_record_flag = False
			self.template_dict[mms]
			self.mms_id = mms
			logger.info(self.mms_id)
			my_rec = AlmaTools(self.key)
			my_rec.get_bib(self.mms_id)
			self.bib_data = my_rec.xml_response_data
			self.parsing_bib_xml()
			if self.template_dict[mms]:
				self.short_record_flag = True
			# print(self.update_flag)
			if self.update_flag or self.short_record_flag:
				my_rec.update_bib(self.mms_id, self.bib_data)
				# print(my_rec.status_code)
				logger.info(self.mms_id + " - updated")
				my_db =DbHandler()
				my_db.db_update_updated(self.mms_id)
			else:
				logger.info("Already has 942 field")



	def get_mms_list_from_alma_set_xlsx_result(self, path_to_xlsx):
		"""Reads 'result.xlsx' file and extracts mms id list from it
		Parameters:
			path_to_xlsx (str) - path to spreadsheet exported from Alma set
		Returns:
			mms_list (list) - contains list of mms ids.
		"""
		mms_list = []
		wb = load_workbook(path_to_xlsx)
		ws= wb.get_sheet_by_name("results")
		for row in ws.iter_rows(min_row=2):
			mms_list.append(row[26].value)
		return(mms_list)



	def custom_update_routine(self, mms_id_list, text_to_change=None, new_text=None):
		"""Manages process of custom updating. Replaces one text with another in alma record
		Parameters:
			mms_id_list(list) - list of mms ids to change
			text_to_change (str) - text in the record you would like to replace.Be carefull, use only the text pattern which is unique on particular record!
			new_text (str) - text which will be inserted instead of 'text_to_change'
		Returns:
			None

		"""
		number_dictionary = {"one":'1',"two":'2',"three":'3',"four":'4',"five":'5',"six":'6',"seven":'7',"eight":'8',"nine":'9', "ten":'10', "zero":'0'}

		self.mms_id_list = mms_id_list
		logger.info("Updating record. Changing '{}' to '{}'".format(text_to_change, new_text))
		sb_mms = None
		if self.key=="sb":
			sb_mms = "9918602951502836"
		self.text_to_change = text_to_change
		self.new_text = new_text
		self.bib_data = None
		for mms in self.mms_id_list:
				flag_to_change_record = False
				if sb_mms:
					logger.info("Sandbox only")
					self.mms = str(sb_mms)
				self.mms_id = mms
				logger.info(self.mms_id)
				my_rec = AlmaTools(self.key)
				my_rec.get_bib(self.mms_id)
				self.bib_data = my_rec.xml_response_data
				if not text_to_change and not new_text:
					# print("here2")
					my_record = parse_xml_to_array(io.StringIO(self.bib_data))[0]
					f245a = my_record["245"]["a"]

					if f245a.startswith("Episode"):
								f490v = f245a.split("-")[0].rstrip(" ")
								my_record["490"]["v"]=f490v
								my_record["800"]["v"]=f490v.lower()
								f245a = "-".join(f245a.split("-")[1:]).lstrip(" ")
								my_record["245"]["a"]=f245a
								self.bib_data = start_xml +str(record_to_xml(my_record)).replace("\\n", "\n").replace("\\", "")+end_xml
								# print(self.bib_data)
								flag_to_change_record = True


				if text_to_change and new_text:
					# print("here1")
					if not new_text in self.bib_data:
						self.bib_data = str(self.bib_data).replace(text_to_change, new_text)
						#print(self.bib_data)
						flag_to_change_record = True
					else:
						logger.info("{} text already in record - nothing to update".format(new_text))
				if flag_to_change_record:
					my_rec.update_bib(self.mms_id, self.bib_data)
					if sb_mms:
						quit()
					if my_rec.status_code == 200:
						logger.info("{} - updated with '{}'".format(self.mms_id, self.bib_data))
					else:
						logger.info(my_rec.status_code)
						logger.info(my_rec.xml_response_data)					
	

	def remove_fields(self, fields):

		record = parse_xml_to_array(io.StringIO(self.bib_data))[0]
		record.remove_fields(fields[0])
		self.bib_data =start_xml+str(record_to_xml(record)).replace("\\n", "\n").replace("\\", "")+end_xml
		

	def general_update_routine(self, key, mms_id,  delete = False, fields = None):

		logger.info("Updating record '{}'".format(mms_id))
		my_rec = AlmaTools(key)
		my_rec.get_bib(mms_id)
		self.bib_data = my_rec.xml_response_data
		if delete:
			self.remove_fields(fields)
		else:
			pass
		# print(my_rec.xml_response_data)
		my_rec.update_bib(mms_id , self.bib_data)
		# print(my_rec.xml_response_data)
		# print(my_rec.status_code)
		

def main():




	mms_list = []
	#change "prod" on "sb" if you require a Sand Box
	my_rec = Manage_fields("prod") 
	#my_rec.cleaning_routine(["9918991865302836"])
	#put your mms ids inside the [] or use  get_mms_list_from_alma_set_xlsx_result
	
	#######################################Example from other types of spreadsheet######################################
	# Example how to use different spreadsheets

	# workook_path = r"Y:\ndha\pre-deposit_prod\LD_working\podcasts\assets\results_Selfie.xlsx"
	# if os.path.exists(workook_path):
	# 	wb = load_workbook(workook_path)
	# 	#Enter name of the working sheet below
	# 	ws= wb.get_sheet_by_name("results")
	# 	#if now headers min_row =1
	# 	for row in ws.iter_rows(min_row=2):
	# 	#21for full results
	# 	#depending on where mms id is row[3] should be changed to number of column started from 0.
	# 		mms = row[26].value
	# 		mms_list.append(row[26].value)
			#print(mms)
			#quit()


	# 	bib_data = my_rec.general_update_routine( "prod", mms,  delete = True, fields = ["520"])
	# ##################################################GET MMSLIST FROM SET####################################################
	# my_alma = AlmaTools("prod")
	# my_alma.get_set_members("",{'limit':'100'})#, "offset":"90"})
	# bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
	# for ind in range(len(bibs)):
	# 	print(bibs[ind])
	# 	mms_list+=[bibs[ind]]
	# print(len(mms_list))
	# print(mms_list)


	###########################Custom update ###############################
	# print(mms_list)
	# my_rec.custom_update_routine(mms_list,'►',' ')
	# my_rec.custom_update_routine(mms_list,'  ',' ')

	#my_rec.custom_update_routine(mms_list,'','')
	#my_rec.custom_update_routine(mms_list)
	
	##############################Getting all mms list from DB ###############################
	# if mms_list == []:
	# 	db_handler = DbHandler()
	# 	my_episodes = db_handler.db_reader(["podcast_name", "mis_mms", "holdings", "item","ie_num"], None, True)
	# 	for episode in my_episodes:
	# 		if "mis_mms" in episode.keys():
	# 			if episode["item"]:
	# 				mms_list.append(episode["mis_mms"])

	#animals matter
	#[Auckland?]
	#[Wellington]
	#Save Animals From Exploitation (Organization : N.Z.),
	#SAFE for Animals (Organisation),

	############################Normal cleaning routine#######################
	# my_rec.cleaning_routine()	# db_handler = DbHandler()
	# my_episodes = db_handler.db_reader(["podcast_name", "mis_mms","episode_title","epis_seas", "epis_numb"], ["Animal matters"], True)
	#my_alma = AlmaTools("prod")
	# my_alma.get_set_members("10310584880002836",{'limit':'100'})#,"offset":"90"})
	# bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
	# for ind in range(len(bibs)):
	# 	print(bibs[ind])
	# 	my_alma.get_bib(bibs[ind])
	# 	title = re.findall(r'<title>(.*?)</title>', my_alma.xml_response_data)[0]
	# 	for epis in my_episodes:
	# 		#print(":".join(epis["episode_title"].split(": ")[1:]).rstrip(", Williams, Guy,").rstrip("."))
	# 		print(epis["episode_title"])
	# 		print(title.rstrip("."))
	# 		print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
	# 		#if ":" in epis["episode_title"] and ":".join(epis["episode_title"].split(": ")[1:]).rstrip("., Williams, Guy,").rstrip(".") == title.rstrip(".") and epis["epis_numb"]:
	# 		if epis["episode_title"].rstrip(".") == title.rstrip(".") and epis["epis_numb"]:
	# 			print("here!!!!!!!!!!")
	# 			my_rec = parse_xml_to_array(io.StringIO(my_alma.xml_response_data))[0]
	# 			# my_rec['490']['v'] = f"S{epis['epis_seas']}:E{epis['epis_numb']}"
	# 			# my_rec['800']['v'] = f"S{epis['epis_seas']}:E{epis['epis_numb']}"
	# 			my_rec['490']['v'] = f"Episode {epis['epis_numb']}"
	# 			my_rec['800']['v'] = f"episode {epis['epis_numb']}"
	# 			bib_data = start_xml +str(record_to_xml(my_rec)).replace("\\n", "\n").replace("\\", "")+end_xml
	# 			my_alma.update_bib(bibs[ind],bib_data)
	################################
	#prop ac 10036805200002836
	#an matter 10310584880002836
	#self refl "10310518530002836"
	#
	#sets = ["10348089760002836", "10315857720002836","10036805200002836","10310621450002836","10310584880002836","10310550350002836","10310518530002836", "10265912860002836"]
		
	##################################################Property academy cleaning#########################################################################
	# lst = ["63","64","65","66","70","71","97","99","101"]
	# sets = ["11790601660002836"]
	# # mms_ids_to_delete= []
	# # my_titles = []
	# # for st in sets:
	# # 	offset_step = 90
	# # 	offset = 0
	# # 	for i in range(20):
	# # 		offset = offset+offset_step
	# my_alma.get_set_members(sets[0],{'limit':'100'})

	# bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	

	# for ind in range(len(bibs)):
	# 		print(ind)
	# 	# print(bibs[ind])
	# 		my_alma.get_bib(bibs[ind])
						
	# 	# if "⎜" in my_alma.xml_response_data:
	# 		my_rec = parse_xml_to_array(io.StringIO(my_alma.xml_response_data))[0]
	# 		# f245a = my_rec["245"]["a"].split("⎜")[0]
	# 		f490v = my_rec["490"]["v"].replace("Episode Episode Episode","Episode").replace("Episode Episode","Episode")#.split("⎜")[-1]
	# 		# if f490v.endswith("."):
	# 		# 	f490v = f490v.rstrip(".")
	# 		f830v = f490v.lower()+"."
	# 		#my_rec["245"]["a"] = f245a
	# 		my_rec["490"]["v"] = f490v
	# 		my_rec["800"]["v"] = f830v

	# 		bib_data = start_xml +str(record_to_xml(my_rec)).replace("\\n", "\n").replace("\\", "")+end_xml
	# 		my_alma.update_bib(bibs[ind], bib_data)
	# 		print(bibs[ind]," - updated")
	# ###########################################################EPIC PODCAST Cleaning#################################################################################


	# sets = ["10315857950002836"]
	# for st in sets:
	# 	offset = 0
	# 	offset_step = 90
	# 	for i in range(3):
			
	# 		print(i)
	# 		print(offset)
	# 		my_alma.get_set_members(st,{'limit':'100',"offset":str(offset)})
	# 		bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
	# 		for ind in range(len(bibs)):
	# 			print(bibs[ind])
	# 			print("here")
	# 			my_alma.get_bib(bibs[ind])
	# 			my_rec = parse_xml_to_array(io.StringIO(my_alma.xml_response_data))[0]
	# 			f490v = my_rec["490"]["v"]
	# 			if len(f490v) == 1 and f490v.isdigit():
	# 				f490v = "Episode "+f490v
	# 				f830v = f490v.lower() + "."
	# 			elif " " in f490v:
	# 				f490v = f490v.replace(" ",":")
	# 				f830v = f490v+ "."

	# 			my_rec["490"]["v"] = f490v
	# 			my_rec["830"]["v"] = f830v
	# 			print(my_rec)
	# 			bib_data = start_xml +str(record_to_xml(my_rec)).replace("\\n", "\n").replace("\\", "")+end_xml
	# 			my_alma.update_bib(bibs[ind], bib_data)
	# 			print("updated")
	# 			offset = offset+offset_step
	# ###########################################################Angus Dunn Cleaning#################################################################################


	# sets = ["10315857950002836"]
	# for st in sets:
	# 	offset = 0
	# 	offset_step = 90
	# 	title_list = []
	# 	for i in range(3):
			
	# 		print(i)
	# 		print(offset)
	# 		my_alma.get_set_members(st,{'limit':'100',"offset":str(offset)})
	# 		bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
	# 		for ind in range(len(bibs)):
	# 			print(bibs[ind])
	# 			my_alma.get_bib(bibs[ind])
	# 			my_rec = parse_xml_to_array(io.StringIO(my_alma.xml_response_data))[0]
	# 			#print(my_rec)

	# 			title_list.append(my_rec["245"]["a"].rstrip("."))				
	# 			if "The Angus Dunn Podcast " in my_alma.xml_response_data:
	# 				f245a = my_rec["245"]["a"].lstrip("The Angus Dunn Podcast ")
	# 				if "-" in f245a:
	# 					divider = "-"
	# 				if ":" in f245a:
	# 					divider = ":"
	# 				f490v =f245a.split(divider)[0]
	# 				f830v = f490v.lower()+"."
	# 				f245a = divider.join(f245a.split(divider)[1:])
	# 				my_rec["245"]["a"] = f245a
	# 				my_rec["490"]["v"] = f490v
	# 				my_rec["800"]["v"] = f830v
	# 				print(my_rec)
	# 				bib_data = start_xml +str(record_to_xml(my_rec)).replace("\\n", "\n").replace("\\", "")+end_xml
	# 				my_alma.update_bib(bibs[ind], bib_data)
	# 				print(bibs[ind]," - updated")
	# 			offset = offset+offset_step
	# print(title_list)
	
##############################################################Animals Matter#########################################
	# title_list = []
	# titles_list = []
	# sets = ["10324342380002836"]
	# #my_titles = ['Best Of Bhuja - May 13 2016', 'Fri, 13 May 2016 17:26:58 +0000', 'May 13 2016', 'Best Of Bhuja'], ['Best Of Bhuja - May 13 2016', 'Fri, 13 May 2016 05:26:58 +0000', 'May 13 2016', 'Best Of Bhuja'], ['Best Of Bhuja - April 29 2016', 'Fri, 29 Apr 2016 15:38:48 +0000', 'April 29 2016', 'Best Of Bhuja'], ['Best Of Bhuja - April 29 2016', 'Fri, 29 Apr 2016 03:38:48 +0000', 'April 29 2016', 'Best Of Bhuja'], ['Best Of Bhuja - April 22 2016', 'Fri, 22 Apr 2016 15:44:48 +0000', 'April 22 2016', 'Best Of Bhuja'], ['Best Of Bhuja - April 22 2016', 'Fri, 22 Apr 2016 03:44:48 +0000', 'April 22 2016', 'Best Of Bhuja'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 18 2016', 'Fri, 18 Mar 2016 18:41:04 +0000', 'March 18 2016', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 18 2016', 'Fri, 18 Mar 2016 05:41:04 +0000', 'March 18 2016', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 11 2016', 'Fri, 11 Mar 2016 15:55:01 +0000', 'March 11 2016', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 11 2016', 'Fri, 11 Mar 2016 02:55:01 +0000', 'March 11 2016', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - February 26 2016', 'Fri, 26 Feb 2016 14:01:38 +0000', 'February 26 2016', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - February 26 2016', 'Fri, 26 Feb 2016 01:01:38 +0000', 'February 26 2016', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - February 19 2016', 'Fri, 19 Feb 2016 12:41:41 +0000', 'February 19 2016', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 18 2015', 'Fri, 18 Dec 2015 16:37:15 +0000', 'December 18 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 18 2015', 'Fri, 18 Dec 2015 03:37:15 +0000', 'December 18 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 11 2015', 'Fri, 11 Dec 2015 15:08:09 +0000', 'December 11 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 11 2015', 'Fri, 11 Dec 2015 02:08:09 +0000', 'December 11 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - November 27 2015', 'Fri, 27 Nov 2015 14:19:53 +0000', 'November 27 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - November 27 2015', 'Fri, 27 Nov 2015 01:19:53 +0000', 'November 27 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - November 20 2015', 'Fri, 20 Nov 2015 19:51:28 +0000', 'November 20 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - September 18 2015', 'Fri, 18 Sep 2015 19:10:20 +0000', 'September 18 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - September 18 2015', 'Fri, 18 Sep 2015 07:10:20 +0000', 'September 18 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 14 2015', 'Fri, 14 Aug 2015 17:36:11 +0000', 'August 14 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 14 2015', 'Fri, 14 Aug 2015 05:36:11 +0000', 'August 14 2015', 'Best Of Bhuja With Leigh Hart & Jason Hoyte']
	# my_new_titles = [['Bhuja to ya!', 'Sun, 03 Nov 2019 19:48:16 +0000'], ['The Bhuja 500, a Wooden Thing & The most musical interview of all time.', 'Sun, 20 Oct 2019 23:04:23 +0000'], ['BREAKING: Sports News, Steve Hansen & Blowies', 'Sun, 13 Oct 2019 21:06:42 +0000'], ["Manfred Mann's Earth Band, Frank Bunce's World Cup thoughts & Agri Chat", 'Mon, 16 Sep 2019 00:58:12 +0000'], ['Snakes, Artificial Politicians & Breaking News', 'Sun, 08 Sep 2019 22:37:56 +0000'], ['Real Estate, Tommy Tippee & Peter Peterson', 'Sun, 01 Sep 2019 04:19:30 +0000'], ["Bananas 'n Piranhas", 'Thu, 29 Aug 2019 04:50:19 +0000'], ['Corona Lomu, Beerfest Faumuina & Rob Mulgoon', 'Mon, 26 Aug 2019 02:51:18 +0000'], ['Herbs, Book Case Solutions & Big Dick Devine', 'Sun, 18 Aug 2019 03:49:36 +0000'], ['Best of Bhuja - Warriors, Breakers & Dime Scooters', 'Fri, 09 Aug 2019 06:05:48 +0000'], ['Sam Cane, Scotty J Stevenson, but mostly Sam Cane', 'Thu, 01 Aug 2019 21:13:39 +0000'], ['Tyla Nathan-Wong, Bernadine Oliver-Kerby & Leigh Matt-Manaia', 'Thu, 25 Jul 2019 05:02:24 +0000'], ['Sir Paul McCartney, Tide Reports and such', 'Thu, 18 Jul 2019 23:58:53 +0000'], ['Villainy, Urzila Carlson & Angelina Grey', 'Mon, 15 Jul 2019 23:01:01 +0000'], ['Fishing & Adventure, Birthdays & Cats', 'Wed, 10 Jul 2019 23:13:52 +0000'], ['Black Seeds, Nude Neighbours & The Moon Landing', 'Sat, 06 Jul 2019 22:01:21 +0000'], ['All Blacks Jerseys, HomeOwners & TXT chat', 'Thu, 04 Jul 2019 03:32:02 +0000'], ["Wah Wahs, Levi Sherwood & Dion Nash's First Cricket Bat", 'Fri, 28 Jun 2019 23:53:26 +0000'], ['Kieran Read, Intermittent Fasting, & Breast Milk Frothers', 'Wed, 26 Jun 2019 05:12:41 +0000'], ['Dion Nash, Marmalade & Spewber', 'Mon, 24 Jun 2019 03:49:20 +0000'], ['Ice Coffee, Floaters & Treasure Island', 'Thu, 20 Jun 2019 00:13:21 +0000'], ["Warriors, Re-Siphoning, & Leigh's Flu-Game", 'Fri, 14 Jun 2019 04:31:47 +0000'], ['Surviving the Titanic, Laser Eyes & Ranch Slider Dressing', 'Fri, 07 Jun 2019 21:36:31 +0000'], ['Grant Elliott, Everest & SportsCafe?', 'Fri, 31 May 2019 04:09:14 +0000'], ['Special Episode: Uncut Jimmy Barnes Interview', 'Fri, 24 May 2019 03:25:54 +0000'], ['Jimmy Barnes, Broccoli Salad & The Nutri-Toilet', 'Fri, 24 May 2019 01:02:32 +0000'], ['Ed Byrne, Ross Noble & Gonorrhea', 'Fri, 17 May 2019 06:39:00 +0000'], ['The Rotunda of Love', 'Fri, 10 May 2019 19:58:05 +0000'], ["It's Back!", 'Fri, 03 May 2019 04:15:53 +0000'], ["Mar 14 -  I'm steamed Bob!", 'Thu, 14 Mar 2019 05:35:49 +0000'], ['Mar 13 - Fast shows a Good show', 'Wed, 13 Mar 2019 05:31:14 +0000'], ['Mar 12 - For Sale - A Montego Viewmaster Trident Sedan deville SL the X Series', 'Tue, 12 Mar 2019 05:53:42 +0000'], ['Mar 11 - Too many Yokos', 'Mon, 11 Mar 2019 05:37:54 +0000'], ["Mar 8 2019 - It's the end of Bhuja as we know it", 'Fri, 08 Mar 2019 03:28:17 +0000'], ['Mar 7 - WHAT HAVE YOU HEARD???', 'Thu, 07 Mar 2019 05:51:36 +0000'], ['Hologram Hoyte', 'Wed, 06 Mar 2019 05:24:45 +0000'], ['How did you get your nudes?', 'Tue, 05 Mar 2019 05:46:29 +0000'], ['Romantic or Erotica', 'Mon, 04 Mar 2019 06:02:49 +0000'], ['Have you got any anusanus?', 'Fri, 01 Mar 2019 06:06:59 +0000'], ['Feb 28 - Go fund yourself', 'Thu, 28 Feb 2019 05:43:44 +0000'], ['Size issues', 'Wed, 27 Feb 2019 05:38:59 +0000'], ['Feb 26 - Text Bombs, Email Chains & Wilson Dixon', 'Tue, 26 Feb 2019 05:59:36 +0000'], ['Not bad for a Mondee', 'Mon, 25 Feb 2019 06:29:17 +0000'], ['Best Of Bhuja - Feb 21 2019', 'Thu, 21 Feb 2019 05:22:15 +0000'], ['HE WENT THAT WAY!', 'Wed, 20 Feb 2019 05:30:33 +0000'], ['Must be weird to pat.', 'Tue, 19 Feb 2019 05:32:28 +0000'], ['Pre-Season Radio', 'Mon, 18 Feb 2019 05:34:52 +0000'], ['Call me Papi', 'Fri, 15 Feb 2019 05:36:18 +0000'], ['The Kaimai Brill Special', 'Wed, 13 Feb 2019 06:58:57 +0000'], ["Jase's Love Salad", 'Tue, 12 Feb 2019 05:48:38 +0000'], ['Bhuja BAFTA and Grammy Special', 'Mon, 11 Feb 2019 05:32:29 +0000'], ['Leonardo da Bhuja', 'Fri, 08 Feb 2019 05:31:04 +0000'], ['Car troubles?', 'Thu, 07 Feb 2019 06:26:45 +0000'], ['God Defend Bhuja', 'Tue, 05 Feb 2019 06:20:44 +0000'], ["Leigh's solo music career, Air NZ safety videos & Expedition Earth", 'Mon, 04 Feb 2019 05:08:12 +0000'], ['Bhuja - Award Winning Radio?', 'Fri, 01 Feb 2019 05:20:21 +0000'], ['Come fly with me.', 'Thu, 31 Jan 2019 06:19:23 +0000'], ['The Bhuja Rumour Mill', 'Wed, 30 Jan 2019 06:36:12 +0000'], ['Six-Degrees Of Separation & Supermarket Sherpa', 'Tue, 29 Jan 2019 19:52:38 +0000'], ['Black Clash, Florence + The Machine & HeathWave', 'Tue, 29 Jan 2019 19:51:41 +0000'], ['Bhuja Live From Orana Park', 'Fri, 25 Jan 2019 02:16:05 +0000'], ['Lime Scooters & Sex Toys', 'Thu, 24 Jan 2019 01:22:29 +0000'], ['The boys back are in town!', 'Wed, 23 Jan 2019 00:56:05 +0000'], ['2018 Clip Show', 'Wed, 12 Dec 2018 03:29:17 +0000'], ['The Final Free For All Friday', 'Wed, 05 Dec 2018 23:39:22 +0000'], ['Bhuja Vinaka!!!', 'Thu, 29 Nov 2018 06:15:08 +0000'], ['Best of Bhuja: Headlamp in the Bedroom', 'Thu, 22 Nov 2018 20:18:41 +0000'], ['Best of Bhuja: Two for Tuesdee', 'Fri, 16 Nov 2018 03:44:52 +0000'], ['The Night Wolf Returns!!', 'Fri, 09 Nov 2018 05:00:07 +0000'], ['Best of Bhuja: Lava Love and Mr. Cheesy', 'Thu, 01 Nov 2018 06:02:12 +0000'], ['Best of Bhuja: Ban It!', 'Fri, 26 Oct 2018 05:59:21 +0000'], ["Best of Bhuja: Uncle Jimrod and the Slice o' Life Forey", 'Thu, 18 Oct 2018 08:07:06 +0000'], ['Best of Bhuja: Rachel Hunter, Anika Moa and the F Bomb', 'Fri, 05 Oct 2018 04:24:26 +0000'], ['The KFC-Adilla', 'Fri, 28 Sep 2018 02:58:48 +0000'], ['Cleaning The Fat Kids & Driving Shoes', 'Fri, 21 Sep 2018 01:52:22 +0000'], ['Bhuja Podcast: Hash pipes, Top Tips and Weeze In Jars', 'Fri, 14 Sep 2018 04:57:05 +0000'], ["Best of Bhuja: Peter Lacey You're An Arsehole Mate", 'Sat, 08 Sep 2018 05:38:28 +0000'], ['Best Of Bhuja: Orange is the new Bhuja', 'Sun, 02 Sep 2018 06:12:18 +0000'], ['Best Of Bhuja: The podcast - Bitcoin, S**tcoin and Couchcoin', 'Sat, 25 Aug 2018 08:13:01 +0000'], ['Best Of Bhuja: Mills and Bhuja - Captive Love', 'Thu, 16 Aug 2018 23:05:39 +0000'], ['Best of Bhuja: Holden Sphincters and Viagra Thursdee', 'Fri, 10 Aug 2018 06:38:54 +0000'], ['Barred Up Tides & the David Bain Tour', 'Fri, 03 Aug 2018 06:37:42 +0000'], ['Best of Bhuja: Barred Up Tides and the David Bain Tour', 'Fri, 03 Aug 2018 06:37:42 +0000'], ["Leigh's Ski Holiday, Big Show Admin & Tom's Ass", 'Sat, 28 Jul 2018 00:01:22 +0000'], ["Best of Bhuja: Leigh's Ski Holiday, big show admin and Tom's ass", 'Sat, 28 Jul 2018 00:01:22 +0000'], ['Massive Heads, Red Wine Top Tips & Leigh Turns 35ish...', 'Fri, 20 Jul 2018 06:15:02 +0000'], ['Best Of Bhuja: Massive heads, red wine top tips and Leigh turns 35...ish', 'Fri, 20 Jul 2018 06:15:02 +0000'], ["Best Of Bhuja: Internet porn, Hoyte's love making and gay funerals", 'Mon, 16 Jul 2018 00:30:59 +0000'], ["Best Of Bhuja: Internet porn, Hoyte's love making and gay funerals", 'Mon, 16 Jul 2018 00:30:59 +0000'], ['Best of Bhuja - The infamous Amy Shark Interview', 'Fri, 06 Jul 2018 06:46:44 +0000'], ['Best of Bhuja: The Podcast - The infamous Amy Shark Interview', 'Fri, 06 Jul 2018 06:46:44 +0000'], ["Best of Bhuja: Bhuja's Got Talent", 'Fri, 29 Jun 2018 07:34:14 +0000'], ["Best of Bhuja: Bhuja's Got Talent", 'Fri, 29 Jun 2018 07:34:14 +0000'], ["Best of Bhuja - Jacinda's Placenta and the F-Bomb", 'Sat, 23 Jun 2018 05:57:28 +0000'], ["Best of Bhuja - Jacinda's Placenta and the F-Bomb", 'Sat, 23 Jun 2018 05:57:28 +0000'], ['Best of Bhuja - Week Two Of The Radio Hauraki Brewery Tour', 'Sat, 16 Jun 2018 03:08:35 +0000'], ['Best of Bhuja - Week Two Of The Radio Hauraki Brewery Tour', 'Sat, 16 Jun 2018 03:08:35 +0000'], ['Best of Bhuja - Week One of the Radio Hauraki Brewery Tour', 'Sat, 09 Jun 2018 01:47:50 +0000'], ['Best of Bhuja: The Hauraki Brewery Tour Week One', 'Sat, 09 Jun 2018 01:47:50 +0000'], ['Best Of Bhuja - The Trivago Lady & Bio Mag G Strings', 'Sat, 02 Jun 2018 01:17:36 +0000'], ['Best Of Bhuja - The Trivago Lady & Bio Mag G Strings', 'Sat, 02 Jun 2018 01:17:36 +0000'], ['Between The Bhujas - Sam Cutler', 'Wed, 30 May 2018 04:17:22 +0000'], ['Between The Bhujas - Sam Cutler', 'Wed, 30 May 2018 04:17:22 +0000'], ['Best of Bhuja - Love Making Top Tips, Royal Wedding & Peter Williams', 'Fri, 25 May 2018 08:33:23 +0000'], ['Best of Bhuja - Love Making Top Tips, Royal Wedding & Peter Williams', 'Fri, 25 May 2018 08:33:23 +0000'], ['Best Of Bhuja - Talking listeners off the ledge & The Rolling Stones', 'Sun, 20 May 2018 01:33:58 +0000'], ['Best Of Bhuja - Talking listeners off the ledge & The Rolling Stones', 'Sun, 20 May 2018 01:33:58 +0000'], ['Best of Bhuja - Bhuja the Political Party & Bumper Stickers', 'Fri, 11 May 2018 05:53:00 +0000'], ['Best of Bhuja - Bhuja the Political Party & Bumper Stickers', 'Fri, 11 May 2018 05:53:00 +0000'], ['Best Of Bhuja - Royal Blood & Mr Chang', 'Sat, 05 May 2018 02:19:28 +0000'], ['Best Of Bhuja - Royal Blood & Mr Chang', 'Sat, 05 May 2018 02:19:28 +0000'], ["Best Of Bhuja - Casting Leigh's Biopic & Bible Chat", 'Sat, 21 Apr 2018 01:44:17 +0000'], ["Best Of Bhuja - Casting Leigh's Biopic & Bible Chat", 'Sat, 21 Apr 2018 01:44:17 +0000'], ['Best of Bhuja - Karl Urban & Hartwatch', 'Sat, 14 Apr 2018 04:38:43 +0000'], ['Best of Bhuja - Karl Urban & Hartwatch', 'Sat, 14 Apr 2018 04:38:43 +0000'], ["Best Of Bhuja - Vance Joy & Jase's Brain Aneurysms", 'Fri, 30 Mar 2018 01:23:38 +0000'], ["Best Of Bhuja - Vance Joy & Jase's Brain Aneurysms", 'Fri, 30 Mar 2018 01:23:38 +0000'], ["Best of Bhuja - Jase's mysterious past & Leigh tries dramatic acting", 'Fri, 23 Mar 2018 22:13:02 +0000'], ["Best of Bhuja - Jase's mysterious past & Leigh tries dramatic acting", 'Fri, 23 Mar 2018 22:13:02 +0000'], ["Best of Bhuja - Chopper Read, A New Host & Jase's Sex Ed", 'Fri, 16 Mar 2018 05:33:12 +0000'], ["Best of Bhuja - Chopper Read, A New Host & Jase's Sex Ed", 'Fri, 16 Mar 2018 05:33:12 +0000'], ['Best Of Bhuja - The Oscars, The Census & Dunedin', 'Sat, 10 Mar 2018 21:48:30 +0000'], ['Best Of Bhuja - The Oscars, The Census & Dunedin', 'Sat, 10 Mar 2018 21:48:30 +0000'], ["Best Of Bhuja - Leigh's Great Aunty Fellatio & Jase's Lovemaking Rib Bib", 'Sun, 04 Mar 2018 02:37:24 +0000'], ["Best Of Bhuja - Leigh's Great Aunty Fellatio & Jase's Lovemaking Rib Bib", 'Sun, 04 Mar 2018 02:37:24 +0000'], ['Best of Bhuja - Faecal Guards, Poos In Pools, Sex Shops & Free For All Friday', 'Sat, 24 Feb 2018 21:21:25 +0000'], ['Best of Bhuja - Faecal Guards, Poos In Pools, Sex Shops & Free For All Friday', 'Sat, 24 Feb 2018 21:21:25 +0000'], ["Best of Bhuja - Leigh's Big Announcement, Winter Olympics & Cremations", 'Sun, 18 Feb 2018 00:27:46 +0000'], ["Best of Bhuja - Leigh's Big Announcement, Winter Olympics & Cremations", 'Sun, 18 Feb 2018 00:27:46 +0000'], ['Best of Bhuja - Waitangi Day, Sex Toys & Babysitters!', 'Fri, 09 Feb 2018 18:56:38 +0000'], ['Best of Bhuja - Waitangi Day, Sex Toys & Babsitters!', 'Fri, 09 Feb 2018 18:56:38 +0000'], ['Best of Bhuja - Parkour, Psychics & Top Tips', 'Fri, 02 Feb 2018 19:29:39 +0000'], ['Best of Bhuja - Parkour, Psychics & Top Tips', 'Fri, 02 Feb 2018 19:29:39 +0000'], ['Best of Bhuja - Pranking Jase, Taylor Swift & Top Tips', 'Sat, 27 Jan 2018 15:16:44 +0000'], ['Best of Bhuja - Pranking Jase, Taylor Swift & Top Tips', 'Sat, 27 Jan 2018 02:16:44 +0000'], ['Best of Bhuja - Chinese weddings, stings & Hillary Barry', 'Sat, 20 Jan 2018 10:31:46 +0000'], ['Best of Bhuja - Chinese weddings, stings & Hillary Barry', 'Fri, 19 Jan 2018 21:31:46 +0000'], ["Best of Bhuja - Christmas parties, Leigh's two left feet & Santa", 'Sun, 10 Dec 2017 09:42:39 +0000'], ["Best of Bhuja - Christmas parties, Leigh's two left feet & Santa", 'Sat, 09 Dec 2017 20:42:39 +0000'], ['Best of Bhuja - Liam Dann, Jesus & Shout Outs', 'Sun, 03 Dec 2017 10:47:13 +0000'], ['Best of Bhuja - Liam Dann, Jesus & Shout Outs', 'Sat, 02 Dec 2017 21:47:13 +0000'], ['Best of Bhuja - Ross Kemp & Dick Pics', 'Mon, 27 Nov 2017 13:12:57 +0000'], ['Best of Bhuja - Ross Kemp & Dick Pics', 'Mon, 27 Nov 2017 00:12:57 +0000'], ['Best of Bhuja - Antiques Road Show, Naked & Afraid and Neil Finn', 'Fri, 17 Nov 2017 16:55:56 +0000'], ['Best of Bhuja - Antiques Road Show, Naked & Afraid and Neil Finn', 'Fri, 17 Nov 2017 03:55:56 +0000'], ["Best of Bhuja - Shelving food, Jase's virginity & Joe Bennett", 'Fri, 10 Nov 2017 15:10:11 +0000'], ["Best of Bhuja - Shelving food, Jase's virginity & Joe Bennett", 'Fri, 10 Nov 2017 02:10:11 +0000'], ['Best of Bhuja - Magnets, tyre pressure & lesser known All Blacks', 'Sat, 04 Nov 2017 13:15:55 +0000'], ['Best of Bhuja - Magnets, tyre pressure & lesser known All Blacks', 'Sat, 04 Nov 2017 00:15:55 +0000'], ["Best of Bhuja - Speed Sperm and Jase's big feet", 'Sun, 29 Oct 2017 13:46:49 +0000'], ["Best of Bhuja - Speed Sperm and Jase's big feet", 'Sun, 29 Oct 2017 00:46:49 +0000'], ['Best of Bhuja - Fishing in Fiji, breast milk dip & Ali Pugh', 'Sat, 21 Oct 2017 20:25:53 +0000'], ['Best of Bhuja - Fishing in Fiji, breast milk dip & Ali Pugh', 'Sat, 21 Oct 2017 07:25:53 +0000'], ['Best of Bhuja - Devices in the bedroom & foreskins', 'Sat, 07 Oct 2017 15:19:36 +0000'], ['Best of Bhuja - Devices in the bedroom & foreskins', 'Sat, 07 Oct 2017 02:19:36 +0000'], ['Best Of Bhuja - Swinger Parties, Fish Chat & Hugh Hefner', 'Fri, 29 Sep 2017 15:31:20 +0000'], ['Best Of Bhuja - Swinger Parties, Fish Chat & Hugh Hefner', 'Fri, 29 Sep 2017 02:31:20 +0000'], ['Best of Bhuja - Emmys Deaths & Uncle Ling', 'Mon, 25 Sep 2017 08:49:53 +0000'], ['Best of Bhuja - Emmys Deaths & Uncle Ling', 'Sun, 24 Sep 2017 19:49:53 +0000'], ['Best of Bhuja - Bill English, "big bush" & excess skin', 'Sat, 16 Sep 2017 17:49:13 +0000'], ['Best of Bhuja - Bill English, "big bush" & excess skin', 'Sat, 16 Sep 2017 05:49:13 +0000'], ["Best of Bhuja - Scotty J Stevenson, Beaver & Uncle Trimble's Autopsy", 'Sat, 09 Sep 2017 17:08:04 +0000'], ["Best of Bhuja - Scotty J Stevenson, Beaver & Uncle Trimble's Autopsy", 'Sat, 09 Sep 2017 05:08:04 +0000'], ["Best of Bhuja -  Jon Toogood, Cancelling Interviews & Leigh's New Music", 'Fri, 01 Sep 2017 14:02:02 +0000'], ["Best of Bhuja -  Jon Toogood, Cancelling Interviews & Leigh's New Music", 'Fri, 01 Sep 2017 02:02:02 +0000'], ['Best Of Bhuja - Dane Coles, Smuggling Drugs & Favourite Dips', 'Sun, 27 Aug 2017 13:38:37 +0000'], ['Best Of Bhuja - Dane Coles, Smuggling Drugs & Favourite Dips', 'Sun, 27 Aug 2017 01:38:37 +0000'], ['Best of Bhuja - Everest Orgasms & Leader Debates', 'Tue, 22 Aug 2017 15:57:35 +0000'], ['Best of Bhuja - Everest Orgasms & Leader Debates', 'Tue, 22 Aug 2017 03:57:35 +0000'], ["Best of Bhuja - Dark web and 'Pets Pets Pets'", 'Sun, 13 Aug 2017 11:00:08 +0000'], ["Best of Bhuja - Dark web and 'Pets Pets Pets'", 'Sat, 12 Aug 2017 23:00:08 +0000'], ["Best of Bhuja - Leigh's stigmata & what gets on your goat", 'Mon, 07 Aug 2017 08:33:37 +0000'], ["Best of Bhuja - Leigh's stigmata & what gets on your goat", 'Sun, 06 Aug 2017 20:33:37 +0000'], ['Best of Bhuja - Amy Shark and mincing zombies', 'Sun, 30 Jul 2017 11:16:32 +0000'], ['Best of Bhuja - Amy Shark and mincing zombies', 'Sat, 29 Jul 2017 23:16:32 +0000'], ['Best of Bhuja - Mr Chang and Female Orgasms', 'Sat, 15 Jul 2017 15:33:58 +0000'], ['Best of Bhuja - Mr Chang and Female Orgasms', 'Sat, 15 Jul 2017 03:33:58 +0000'], ['Best of Bhuja - Neil deGrasse Tyson', 'Mon, 10 Jul 2017 08:50:42 +0000'], ['Best of Bhuja - Neil deGrasse Tyson', 'Sun, 09 Jul 2017 20:50:42 +0000'], ['Best of Bhuja - Jordan Luck and Peter Burling', 'Mon, 03 Jul 2017 11:02:11 +0000'], ['Best of Bhuja - Jordan Luck and Peter Burling', 'Sun, 02 Jul 2017 23:02:11 +0000'], ['Best of Bhuja - A class drug amnesty bins and school discos', 'Mon, 26 Jun 2017 08:27:09 +0000'], ['Best of Bhuja - A class drug amnesty bins and school discos', 'Sun, 25 Jun 2017 20:27:09 +0000'], ["Best of Bhuja - Jas' dodgy answerphone & Hamilton strip clubs", 'Sat, 17 Jun 2017 16:04:20 +0000'], ["Best of Bhuja - Jas' dodgy answerphone & Hamilton strip clubs", 'Sat, 17 Jun 2017 04:04:20 +0000'], ['Best of Bhuja - Nude cooking & Scotty J Stevenson', 'Sun, 11 Jun 2017 19:58:44 +0000'], ['Best of Bhuja - Nude cooking & Scotty J Stevenson', 'Sun, 11 Jun 2017 07:58:44 +0000'], ['Best of Bhuja - Equipping yachts with torpedoes & putting down cats', 'Sun, 04 Jun 2017 08:57:37 +0000'], ['Best of Bhuja - Equipping yachts with torpedoes & putting down cats', 'Sat, 03 Jun 2017 20:57:37 +0000'], ['Best of Bhuja - Chad Chewbakkaman & Penis Colours', 'Sun, 28 May 2017 11:22:51 +0000'], ['Best of Bhuja - Chad Chewbakkaman & Penis Colours', 'Sat, 27 May 2017 23:22:51 +0000'], ['Best of Bhuja - Joseph Parker & baboon rape', 'Sun, 21 May 2017 19:44:04 +0000'], ['Best of Bhuja - Joseph Parker & baboon rape', 'Sun, 21 May 2017 07:44:04 +0000'], ["Best Of Bhuja - Matt's Head Issue, Celibacy & Guinea Pigs", 'Fri, 12 May 2017 14:20:01 +0000'], ["Best Of Bhuja - Matt's Head Issue, Celibacy & Guinea Pigs", 'Fri, 12 May 2017 02:20:01 +0000'], ['Best Of Bhuja - Fiji Edition', 'Sat, 06 May 2017 18:19:23 +0000'], ['Best Of Bhuja - Fiji Edition', 'Sat, 06 May 2017 06:19:23 +0000'], ['Best Of Bhuja - Cat Sedation, Shelving Vegetables & Sexual Scenarios.', 'Sat, 29 Apr 2017 15:15:24 +0000'], ['Best Of Bhuja - Cat Sedation, Shelving Vegetables & Sexual Scenarios.', 'Sat, 29 Apr 2017 03:15:24 +0000'], ['Best Of Bhuja - Aaron Cruden, The Bachelor, Narcotics & Exorcisms', 'Sun, 09 Apr 2017 08:39:10 +0000'], ['Best Of Bhuja - Aaron Cruden, The Bachelor, Narcotics & Exorcisms', 'Sat, 08 Apr 2017 20:39:10 +0000'], ['Best Of Bhuja - Millen Baird, Jase Gets Sexist & Bennie The Homosexual Beagle', 'Sun, 02 Apr 2017 08:34:52 +0000'], ['Best Of Bhuja - Millen Baird, Jase Gets Sexist & Bennie The Homosexual Beagle', 'Sat, 01 Apr 2017 20:34:52 +0000'], ['Best of Bhuja - Cocksy the Builder, Viking DNA & Happy Endings', 'Mon, 27 Mar 2017 08:46:21 +0000'], ['Best of Bhuja - Cocksy the Builder, Viking DNA & Happy Endings', 'Sun, 26 Mar 2017 19:46:21 +0000'], ["Best Of Bhuja - Jacinda Ardern, Jesus' workload & how to drink Bourbon", 'Sun, 19 Mar 2017 12:11:39 +0000'], ["Best Of Bhuja - Jacinda Ardern, Jesus' workload & how to drink Bourbon", 'Sat, 18 Mar 2017 23:11:39 +0000'], ['Best Of Bhuja - Love in an elevator, Uncle Don & Joe Wheeler from The Highlanders', 'Sun, 12 Mar 2017 12:34:31 +0000'], ['Best Of Bhuja - Love in an elevator, Uncle Don & Joe Wheeler from The Highlanders', 'Sat, 11 Mar 2017 23:34:31 +0000'], ['Best of Bhuja - Elemeno P, Delta Riggs & The Bhuja Financial Report', 'Tue, 07 Mar 2017 17:11:52 +0000'], ['Best of Bhuja - Elemeno P, Delta Riggs & The Bhuja Financial Report', 'Tue, 07 Mar 2017 04:11:52 +0000'], ['Best of Bhuja - E Street Band & Taxi Chittys', 'Tue, 28 Feb 2017 09:59:22 +0000'], ['Best of Bhuja - E Street Band & Taxi Chittys', 'Mon, 27 Feb 2017 20:59:22 +0000'], ["Best of Bhuja - Russian Subs, Uncle Tony's Funeral & Sex Toys", 'Thu, 23 Feb 2017 08:41:28 +0000'], ["Best of Bhuja - Russian Subs, Uncle Tony's Funeral & Sex Toys", 'Wed, 22 Feb 2017 19:41:28 +0000'], ["Best Of Bhuja - Free Passes, Ben Smith's Baby & Drinking Urine", 'Mon, 13 Feb 2017 19:29:27 +0000'], ["Best Of Bhuja - Free Passes, Ben Smith's Baby & Drinking Urine", 'Mon, 13 Feb 2017 06:29:27 +0000'], ['Best of Bhuja - Barry Soper, Fish Chat & Exercise', 'Wed, 08 Feb 2017 08:51:36 +0000'], ['Best of Bhuja - Barry Soper, Fish Chat & Exercise', 'Tue, 07 Feb 2017 19:51:36 +0000'], ["Best of Bhuja - Richie's Wedding & Eco-Lodge Christmas", 'Wed, 01 Feb 2017 09:34:44 +0000'], ["Best of Bhuja - Richie's Wedding & Eco-Lodge Christmas", 'Tue, 31 Jan 2017 20:34:44 +0000'], ['Best of Bhuja - Sex Positions, Lollapabhuja & Star Wars', 'Tue, 20 Dec 2016 16:14:23 +0000'], ['Best of Bhuja - Sex Positions, Lollapabhuja & Star Wars', 'Tue, 20 Dec 2016 03:14:23 +0000'], ['Best of Bhuja  - Joseph Parker and are we good lovers?', 'Tue, 13 Dec 2016 09:03:23 +0000'], ['Best of Bhuja  - Joseph Parker and are we good lovers?', 'Mon, 12 Dec 2016 20:03:23 +0000'], ['Best of Bhuja - Female co-hosts, VHS toastie makers & losing your virginity', 'Mon, 05 Dec 2016 17:11:08 +0000'], ['Best of Bhuja - Female co-hosts, VHS toastie makers & losing your virginity', 'Mon, 05 Dec 2016 04:11:08 +0000'], ['Best of Bhuja - Asparagus wine, circumcisions & the beginning of Radio Hauraki', 'Tue, 29 Nov 2016 08:16:20 +0000'], ['Best of Bhuja - Asparagus wine, circumcisions & the beginning of Radio Hauraki', 'Mon, 28 Nov 2016 19:16:20 +0000'], ['Best Of Bhuja - Brian Tamaki & oversized balls', 'Tue, 22 Nov 2016 08:53:20 +0000'], ['Best Of Bhuja - Brian Tamaki & oversized balls', 'Mon, 21 Nov 2016 19:53:20 +0000'], ['Best Of Bhuja - Hire-A-Hubby, Chris Cornell & Live From Dunedin', 'Mon, 14 Nov 2016 08:05:09 +0000'], ['Best Of Bhuja - Hire-A-Hubby, Chris Cornell & Live From Dunedin', 'Sun, 13 Nov 2016 19:05:09 +0000'], ['Best Of Bhuja - Radio Crossovers, Xmas Party Attire & Nathan McCullum', 'Sat, 05 Nov 2016 14:02:05 +0000'], ['Best Of Bhuja - Radio Crossovers, Xmas Party Attire & Nathan McCullum', 'Sat, 05 Nov 2016 01:02:05 +0000'], ['Best Of Bhuja - Pirate Radio, Massages, Circumsicions & Afternoon Lovers', 'Mon, 31 Oct 2016 15:36:47 +0000'], ['Best Of Bhuja - Pirate Radio, Massages, Circumsicions & Afternoon Lovers', 'Mon, 31 Oct 2016 02:36:47 +0000'], ['Best Of Bhuja - Fat Kids, Kieran Read & The Return Of Curry Night', 'Tue, 25 Oct 2016 15:30:09 +0000'], ['Best Of Bhuja - Fat Kids, Kieran Read & The Return Of Curry Night', 'Tue, 25 Oct 2016 02:30:09 +0000'], ["Best Of Bhuja - Alf From 'Home & Away' & Kebabs", 'Mon, 17 Oct 2016 16:30:42 +0000'], ["Best Of Bhuja - Alf From 'Home & Away' & Kebabs", 'Mon, 17 Oct 2016 03:30:42 +0000'], ['Best Of Bhuja - All Things Bhuja Featuring Big Show Admin', 'Mon, 10 Oct 2016 10:36:41 +0000'], ['Best Of Bhuja - All Things Bhuja Featuring Big Show Admin', 'Sun, 09 Oct 2016 21:36:41 +0000'], ['Best Of Bhuja - Spring Garage Sale And Some Whacky Improv', 'Wed, 28 Sep 2016 09:30:06 +0000'], ['Best Of Bhuja - Spring Garage Sale And Some Whacky Improv', 'Tue, 27 Sep 2016 20:30:06 +0000'], ["Best Of Bhuja - Don't Mix Herbal Ignite With BBQ Ignite", 'Tue, 27 Sep 2016 16:47:23 +0000'], ["Best Of Bhuja - Don't Mix Herbal Ignite With BBQ Ignite", 'Tue, 27 Sep 2016 03:47:23 +0000'], ['Best Of Bhuja - Is This The End Of Curry Night?', 'Thu, 15 Sep 2016 18:24:49 +0000'], ['Best Of Bhuja - Cheese Toastie On A Wednesdee', 'Thu, 15 Sep 2016 08:19:06 +0000'], ['Best Of Bhuja - Is This The End Of Curry Night?', 'Thu, 15 Sep 2016 06:24:49 +0000'], ['Best Of Bhuja - Cheese Toastie On A Wednesdee', 'Wed, 14 Sep 2016 20:19:06 +0000'], ['Best Of Bhuja - KFC Drive-Thru', 'Fri, 09 Sep 2016 13:43:43 +0000'], ['Best Of Bhuja - KFC Drive-Thru', 'Fri, 09 Sep 2016 01:43:43 +0000'], ['Best Of Bhuja - Dog Poos & Fishing', 'Thu, 08 Sep 2016 10:49:13 +0000'], ['Best Of Bhuja - Dog Poos & Fishing', 'Wed, 07 Sep 2016 22:49:13 +0000'], ['Best Of Bhuja - On A Thursdee', 'Wed, 07 Sep 2016 14:43:51 +0000'], ['Best Of Bhuja - On A Thursdee', 'Wed, 07 Sep 2016 02:43:51 +0000'], ['Best Of Bhuja - Mondee Or Was It A Tuesdee', 'Thu, 01 Sep 2016 14:43:36 +0000'], ['Best Of Bhuja - Mondee Or Was It A Tuesdee', 'Thu, 01 Sep 2016 02:43:36 +0000'], ['Best Of The Bhuja Week - August 22 - 26 2016', 'Sat, 27 Aug 2016 18:04:50 +0000'], ['Best Of The Bhuja Week - August 22 - 26 2016', 'Sat, 27 Aug 2016 06:04:50 +0000'], ['Best Of The Bhuja Week - August 15 - 19 2016', 'Sun, 21 Aug 2016 12:05:12 +0000'], ['Best Of The Bhuja Week - August 15 - 19 2016', 'Sun, 21 Aug 2016 00:05:12 +0000'], ['Best Of The Bhuja Week - August 8 - 12 2016', 'Fri, 12 Aug 2016 18:49:25 +0000'], ['Best Of The Bhuja Week - August 8 - 12 2016', 'Fri, 12 Aug 2016 06:49:25 +0000'], ['Best Of The Bhuja Week - June 27-1 July 2016', 'Fri, 01 Jul 2016 16:49:24 +0000'], ['Best Of The Bhuja Week - June 27-1 July 2016', 'Fri, 01 Jul 2016 04:49:24 +0000'], ['Best Of The Bhuja Week - June 20-24 2016', 'Fri, 24 Jun 2016 12:48:01 +0000'], ['Best Of The Bhuja Week - June 20-24 2016', 'Fri, 24 Jun 2016 00:48:01 +0000'], ['Best Of The Bhuja Week - June 13-17 2016', 'Fri, 17 Jun 2016 14:12:10 +0000'], ['Best Of The Bhuja Week - June 13-17 2016', 'Fri, 17 Jun 2016 02:12:10 +0000'], ['Best Of Bhuja - June 10 2016', 'Sun, 12 Jun 2016 13:54:44 +0000'], ['Best Of Bhuja - June 10 2016', 'Sun, 12 Jun 2016 01:54:44 +0000'], ['Best Of Bhuja - June 3 2016', 'Fri, 03 Jun 2016 15:23:16 +0000'], ['Best Of Bhuja - June 3 2016', 'Fri, 03 Jun 2016 03:23:16 +0000'], ['Best Of Bhuja - May 27 2016', 'Sat, 28 May 2016 07:48:30 +0000'], ['Best Of Bhuja - May 27 2016', 'Fri, 27 May 2016 19:48:30 +0000'], ['Best Of Bhuja - May 20 2016', 'Fri, 20 May 2016 15:21:33 +0000'], ['Best Of Bhuja - May 20 2016', 'Fri, 20 May 2016 03:21:33 +0000'], ['Best Of Bhuja - May 13 2016', 'Fri, 13 May 2016 17:26:58 +0000'], ['Best Of Bhuja - May 13 2016', 'Fri, 13 May 2016 05:26:58 +0000'], ['Best Of Bhuja - May 6 2016', 'Sat, 07 May 2016 14:50:57 +0000'], ['Best Of Bhuja - May 6 2016', 'Sat, 07 May 2016 02:50:57 +0000'], ['Best Of Bhuja - April 29 2016', 'Fri, 29 Apr 2016 15:38:48 +0000'], ['Best Of Bhuja - April 29 2016', 'Fri, 29 Apr 2016 03:38:48 +0000'], ['Best Of Bhuja - April 22 2016', 'Fri, 22 Apr 2016 15:44:48 +0000'], ['Best Of Bhuja - April 22 2016', 'Fri, 22 Apr 2016 03:44:48 +0000'], ['Best Of Bhuja - April 15 2016', 'Sat, 16 Apr 2016 14:24:00 +0000'], ['Best Of Bhuja - April 15 2016', 'Sat, 16 Apr 2016 02:24:00 +0000'], ['Best Of Bhuja - April 8 2016', 'Fri, 08 Apr 2016 12:31:53 +0000'], ['Best Of Bhuja - April 8 2016', 'Fri, 08 Apr 2016 00:31:53 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - April 1 2016', 'Fri, 01 Apr 2016 14:19:53 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - April 1 2016', 'Fri, 01 Apr 2016 01:19:53 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 18 2016', 'Fri, 18 Mar 2016 18:41:04 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 18 2016', 'Fri, 18 Mar 2016 05:41:04 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 11 2016', 'Fri, 11 Mar 2016 15:55:01 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 11 2016', 'Fri, 11 Mar 2016 02:55:01 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 4 2016', 'Fri, 04 Mar 2016 15:22:02 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - March 4 2016', 'Fri, 04 Mar 2016 02:22:02 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - February 26 2016', 'Fri, 26 Feb 2016 14:01:38 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - February 26 2016', 'Fri, 26 Feb 2016 01:01:38 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - February 19 2016', 'Fri, 19 Feb 2016 12:41:41 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - February 19 2016', 'Thu, 18 Feb 2016 23:41:41 +0000'], ['Best Of Bhuja - February 12 2016', 'Fri, 12 Feb 2016 17:19:44 +0000'], ['Best Of Bhuja - February 12 2016', 'Fri, 12 Feb 2016 04:19:44 +0000'], ['Best Of Bhuja Podcast - January 29 2016', 'Fri, 29 Jan 2016 17:37:51 +0000'], ['Best Of Bhuja Podcast - January 29 2016', 'Fri, 29 Jan 2016 04:37:51 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - January 22 2016', 'Fri, 22 Jan 2016 14:45:45 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - January 22 2016', 'Fri, 22 Jan 2016 01:45:45 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 18 2015', 'Fri, 18 Dec 2015 16:37:15 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 18 2015', 'Fri, 18 Dec 2015 03:37:15 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 11 2015', 'Fri, 11 Dec 2015 15:08:09 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 11 2015', 'Fri, 11 Dec 2015 02:08:09 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 4 2015', 'Fri, 04 Dec 2015 15:30:29 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - December 4 2015', 'Fri, 04 Dec 2015 02:30:29 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - November 27 2015', 'Fri, 27 Nov 2015 14:19:53 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - November 27 2015', 'Fri, 27 Nov 2015 01:19:53 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - November 20 2015', 'Sat, 21 Nov 2015 08:51:28 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - November 20 2015', 'Fri, 20 Nov 2015 19:51:28 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - October 23 2015', 'Fri, 23 Oct 2015 20:48:33 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - October 23 2015', 'Fri, 23 Oct 2015 07:48:33 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - October 16 2015', 'Sat, 17 Oct 2015 10:44:17 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - October 16 2015', 'Fri, 16 Oct 2015 21:44:17 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - October 9 2015', 'Sat, 10 Oct 2015 10:17:08 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - October 9 2015', 'Fri, 09 Oct 2015 21:17:08 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - October 2 2015', 'Fri, 02 Oct 2015 12:28:42 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - October 2 2015', 'Thu, 01 Oct 2015 23:28:42 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - September 25 2015', 'Fri, 25 Sep 2015 14:49:06 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - September 25 2015', 'Fri, 25 Sep 2015 02:49:06 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - September 18 2015', 'Fri, 18 Sep 2015 19:10:20 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - September 18 2015', 'Fri, 18 Sep 2015 07:10:20 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte – September 11 2015', 'Fri, 11 Sep 2015 17:15:29 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte – September 11 2015', 'Fri, 11 Sep 2015 05:15:29 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte – September 4 2015', 'Fri, 04 Sep 2015 17:49:47 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte – September 4 2015', 'Fri, 04 Sep 2015 05:49:47 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 28 2015', 'Fri, 28 Aug 2015 16:17:05 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 28 2015', 'Fri, 28 Aug 2015 04:17:05 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 21 2015', 'Sat, 22 Aug 2015 15:23:55 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 21 2015', 'Sat, 22 Aug 2015 03:23:55 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 14 2015', 'Fri, 14 Aug 2015 17:36:11 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 14 2015', 'Fri, 14 Aug 2015 05:36:11 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 7 2015', 'Fri, 07 Aug 2015 17:50:47 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - August 7 2015', 'Fri, 07 Aug 2015 05:50:47 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 31 2015', 'Fri, 31 Jul 2015 17:38:27 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 31 2015', 'Fri, 31 Jul 2015 05:38:27 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 24 2015', 'Fri, 24 Jul 2015 15:49:40 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 24 2015', 'Fri, 24 Jul 2015 03:49:40 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 17 2015', 'Fri, 17 Jul 2015 15:00:07 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 17 2015', 'Fri, 17 Jul 2015 03:00:07 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 10 2015', 'Mon, 13 Jul 2015 12:16:34 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 10 2015', 'Mon, 13 Jul 2015 00:16:34 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 3 2015', 'Fri, 03 Jul 2015 17:43:53 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - July 3 2015', 'Fri, 03 Jul 2015 05:43:53 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - June 26 2015', 'Wed, 01 Jul 2015 15:18:36 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - June 19 2015', 'Wed, 01 Jul 2015 15:18:34 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - June 12 2015', 'Wed, 01 Jul 2015 15:18:34 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - June 5 2015', 'Wed, 01 Jul 2015 15:18:33 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 29 2015', 'Wed, 01 Jul 2015 15:18:32 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 22 2015', 'Wed, 01 Jul 2015 15:18:31 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 15 2015', 'Wed, 01 Jul 2015 15:18:30 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 8 2015', 'Wed, 01 Jul 2015 15:18:29 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 1 2015', 'Wed, 01 Jul 2015 15:18:28 +0000'], ['Best Of Bhuja Podcast - 24 April 2015', 'Wed, 01 Jul 2015 15:18:27 +0000'], ['Best Of Bhuja! Podcast - April 2 2015', 'Wed, 01 Jul 2015 15:18:27 +0000'], ['Best Of Bhuja! - March 20 2015', 'Wed, 01 Jul 2015 15:18:26 +0000'], ['Best Of Bhuja - March 13 2015', 'Wed, 01 Jul 2015 15:18:25 +0000'], ['Best Of Bhuja! - March 6 2015', 'Wed, 01 Jul 2015 15:18:24 +0000'], ['Best Of Bhuja - Feb 27 2015', 'Wed, 01 Jul 2015 15:18:23 +0000'], ['Best Of Bhuja - Feb 20 2015', 'Wed, 01 Jul 2015 15:18:21 +0000'], ['Best Of Sports Bhuja - March 29', 'Wed, 01 Jul 2015 15:18:20 +0000'], ['Best Of Sports Bhuja - April 5', 'Wed, 01 Jul 2015 15:18:20 +0000'], ['Best Of Sports Bhuja - April 12', 'Wed, 01 Jul 2015 15:18:19 +0000'], ['Best Of Sports Bhuja - April 19', 'Wed, 01 Jul 2015 15:18:18 +0000'], ['Best Of Sports Bhuja - April 26', 'Wed, 01 Jul 2015 15:18:17 +0000'], ['Best Of Sports Bhuja - May 3', 'Wed, 01 Jul 2015 15:18:17 +0000'], ['Best Of Sports Bhuja - May 10', 'Wed, 01 Jul 2015 15:18:16 +0000'], ['Best Of Sports Bhuja - May 17', 'Wed, 01 Jul 2015 15:18:15 +0000'], ['Best Of Sports Bhuja - May 24', 'Wed, 01 Jul 2015 15:18:14 +0000'], ['Best Of Sports Bhuja - May 31', 'Wed, 01 Jul 2015 15:18:14 +0000'], ['Best Of Sports Bhuja - June 7', 'Wed, 01 Jul 2015 15:18:13 +0000'], ['Best Of Sports Bhuja - June 14', 'Wed, 01 Jul 2015 15:18:12 +0000'], ['Best Of Sports Bhuja - June 21', 'Wed, 01 Jul 2015 15:18:11 +0000'], ['Best Of Sports Bhuja - June 28', 'Wed, 01 Jul 2015 15:18:11 +0000'], ['Best Of Sports Bhuja - July 5', 'Wed, 01 Jul 2015 15:18:10 +0000'], ['Best Of Sports Bhuja - July 12', 'Wed, 01 Jul 2015 15:18:10 +0000'], ['Best Of Sports Bhuja - July 19', 'Wed, 01 Jul 2015 15:18:08 +0000'], ['Best Of Sports Bhuja - July 26', 'Wed, 01 Jul 2015 15:18:08 +0000'], ['Best Of Sports Bhuja - August 2', 'Wed, 01 Jul 2015 15:18:07 +0000'], ['Best Of Sports Bhuja - August 9', 'Wed, 01 Jul 2015 15:18:06 +0000'], ['Best Of Sports Bhuja - August 16', 'Wed, 01 Jul 2015 15:18:05 +0000'], ['Best Of Sports Bhuja - August 23', 'Wed, 01 Jul 2015 15:18:05 +0000'], ['Best Of Sports Bhuja - August 30', 'Wed, 01 Jul 2015 15:18:03 +0000'], ['Best Of Sports Bhuja - September 6', 'Wed, 01 Jul 2015 15:18:01 +0000'], ['Best Of Sports Bhuja - September 13', 'Wed, 01 Jul 2015 15:18:00 +0000'], ['Best Of Sports Bhuja - September 20', 'Wed, 01 Jul 2015 15:18:00 +0000'], ['Best Of Sports Bhuja - September 27', 'Wed, 01 Jul 2015 15:17:59 +0000'], ['Best Of Sports Bhuja - October 4', 'Wed, 01 Jul 2015 15:17:59 +0000'], ['Best Of Sports Bhuja - October 18', 'Wed, 01 Jul 2015 15:17:58 +0000'], ['Best Of Sports Bhuja - October 11', 'Wed, 01 Jul 2015 15:17:58 +0000'], ['Best Of Sports Bhuja - October 25', 'Wed, 01 Jul 2015 15:17:57 +0000'], ['Best Of Sports Bhuja - November 1', 'Wed, 01 Jul 2015 15:17:56 +0000'], ['Best of Sports Bhuja - November 8', 'Wed, 01 Jul 2015 15:17:55 +0000'], ['Best Of Sports Bhuja - November 15', 'Wed, 01 Jul 2015 15:17:54 +0000'], ['Best Of Sports Bhuja - November 22', 'Wed, 01 Jul 2015 15:17:53 +0000'], ['Best Of Sports Bhuja - December 6', 'Wed, 01 Jul 2015 15:17:52 +0000'], ['Best Of Sports Bhuja - November 29', 'Wed, 01 Jul 2015 15:17:52 +0000'], ['Best Of Sports Bhuja - December 13', 'Wed, 01 Jul 2015 15:17:51 +0000'], ['Best Of Sports Bhuja - December 20', 'Wed, 01 Jul 2015 15:17:50 +0000'], ['Best Of Bhuja! - Jan 16 2015', 'Wed, 01 Jul 2015 15:17:49 +0000'], ['Best Of Bhuja! - Jan 23 2015', 'Wed, 01 Jul 2015 15:17:49 +0000'], ['Best Of Bhuja! - Jan 30 2015', 'Wed, 01 Jul 2015 15:17:48 +0000'], ['Best Of Bhuja! - Feb 6 2015', 'Wed, 01 Jul 2015 15:17:47 +0000'], ['Best Of Bhuja! - Feb 13 2015', 'Wed, 01 Jul 2015 15:17:47 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - June 26 2015', 'Wed, 01 Jul 2015 03:18:36 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - June 19 2015', 'Wed, 01 Jul 2015 03:18:34 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - June 12 2015', 'Wed, 01 Jul 2015 03:18:34 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - June 5 2015', 'Wed, 01 Jul 2015 03:18:33 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 29 2015', 'Wed, 01 Jul 2015 03:18:32 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 22 2015', 'Wed, 01 Jul 2015 03:18:31 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 15 2015', 'Wed, 01 Jul 2015 03:18:30 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 8 2015', 'Wed, 01 Jul 2015 03:18:29 +0000'], ['Best Of Bhuja With Leigh Hart & Jason Hoyte - May 1 2015', 'Wed, 01 Jul 2015 03:18:28 +0000'], ['Best Of Bhuja! Podcast - April 2 2015', 'Wed, 01 Jul 2015 03:18:27 +0000'], ['Best Of Bhuja Podcast - 24 April 2015', 'Wed, 01 Jul 2015 03:18:27 +0000'], ['Best Of Bhuja! - March 20 2015', 'Wed, 01 Jul 2015 03:18:26 +0000'], ['Best Of Bhuja - March 13 2015', 'Wed, 01 Jul 2015 03:18:25 +0000'], ['Best Of Bhuja! - March 6 2015', 'Wed, 01 Jul 2015 03:18:24 +0000'], ['Best Of Bhuja - Feb 27 2015', 'Wed, 01 Jul 2015 03:18:23 +0000'], ['Best Of Bhuja - Feb 20 2015', 'Wed, 01 Jul 2015 03:18:21 +0000'], ['Best Of Sports Bhuja - March 29', 'Wed, 01 Jul 2015 03:18:20 +0000'], ['Best Of Sports Bhuja - April 5', 'Wed, 01 Jul 2015 03:18:20 +0000'], ['Best Of Sports Bhuja - April 12', 'Wed, 01 Jul 2015 03:18:19 +0000'], ['Best Of Sports Bhuja - April 19', 'Wed, 01 Jul 2015 03:18:18 +0000'], ['Best Of Sports Bhuja - April 26', 'Wed, 01 Jul 2015 03:18:17 +0000'], ['Best Of Sports Bhuja - May 3', 'Wed, 01 Jul 2015 03:18:17 +0000'], ['Best Of Sports Bhuja - May 10', 'Wed, 01 Jul 2015 03:18:16 +0000'], ['Best Of Sports Bhuja - May 17', 'Wed, 01 Jul 2015 03:18:15 +0000'], ['Best Of Sports Bhuja - May 31', 'Wed, 01 Jul 2015 03:18:14 +0000'], ['Best Of Sports Bhuja - May 24', 'Wed, 01 Jul 2015 03:18:14 +0000'], ['Best Of Sports Bhuja - June 7', 'Wed, 01 Jul 2015 03:18:13 +0000'], ['Best Of Sports Bhuja - June 14', 'Wed, 01 Jul 2015 03:18:12 +0000'], ['Best Of Sports Bhuja - June 21', 'Wed, 01 Jul 2015 03:18:11 +0000'], ['Best Of Sports Bhuja - June 28', 'Wed, 01 Jul 2015 03:18:11 +0000'], ['Best Of Sports Bhuja - July 5', 'Wed, 01 Jul 2015 03:18:10 +0000'], ['Best Of Sports Bhuja - July 12', 'Wed, 01 Jul 2015 03:18:10 +0000'], ['Best Of Sports Bhuja - July 19', 'Wed, 01 Jul 2015 03:18:08 +0000'], ['Best Of Sports Bhuja - July 26', 'Wed, 01 Jul 2015 03:18:08 +0000'], ['Best Of Sports Bhuja - August 2', 'Wed, 01 Jul 2015 03:18:07 +0000'], ['Best Of Sports Bhuja - August 9', 'Wed, 01 Jul 2015 03:18:06 +0000'], ['Best Of Sports Bhuja - August 16', 'Wed, 01 Jul 2015 03:18:05 +0000'], ['Best Of Sports Bhuja - August 23', 'Wed, 01 Jul 2015 03:18:05 +0000'], ['Best Of Sports Bhuja - August 30', 'Wed, 01 Jul 2015 03:18:03 +0000'], ['Best Of Sports Bhuja - September 6', 'Wed, 01 Jul 2015 03:18:01 +0000'], ['Best Of Sports Bhuja - September 13', 'Wed, 01 Jul 2015 03:18:00 +0000'], ['Best Of Sports Bhuja - September 20', 'Wed, 01 Jul 2015 03:18:00 +0000'], ['Best Of Sports Bhuja - September 27', 'Wed, 01 Jul 2015 03:17:59 +0000'], ['Best Of Sports Bhuja - October 4', 'Wed, 01 Jul 2015 03:17:59 +0000'], ['Best Of Sports Bhuja - October 18', 'Wed, 01 Jul 2015 03:17:58 +0000'], ['Best Of Sports Bhuja - October 11', 'Wed, 01 Jul 2015 03:17:58 +0000'], ['Best Of Sports Bhuja - October 25', 'Wed, 01 Jul 2015 03:17:57 +0000'], ['Best Of Sports Bhuja - November 1', 'Wed, 01 Jul 2015 03:17:56 +0000'], ['Best of Sports Bhuja - November 8', 'Wed, 01 Jul 2015 03:17:55 +0000'], ['Best Of Sports Bhuja - November 15', 'Wed, 01 Jul 2015 03:17:54 +0000'], ['Best Of Sports Bhuja - November 22', 'Wed, 01 Jul 2015 03:17:53 +0000'], ['Best Of Sports Bhuja - November 29', 'Wed, 01 Jul 2015 03:17:52 +0000'], ['Best Of Sports Bhuja - December 6', 'Wed, 01 Jul 2015 03:17:52 +0000'], ['Best Of Sports Bhuja - December 13', 'Wed, 01 Jul 2015 03:17:51 +0000'], ['Best Of Sports Bhuja - December 20', 'Wed, 01 Jul 2015 03:17:50 +0000'], ['Best Of Bhuja! - Jan 16 2015', 'Wed, 01 Jul 2015 03:17:49 +0000'], ['Best Of Bhuja! - Jan 23 2015', 'Wed, 01 Jul 2015 03:17:49 +0000'], ['Best Of Bhuja! - Jan 30 2015', 'Wed, 01 Jul 2015 03:17:48 +0000'], ['Best Of Bhuja! - Feb 6 2015', 'Wed, 01 Jul 2015 03:17:47 +0000'], ['Best Of Bhuja! - Feb 13 2015', 'Wed, 01 Jul 2015 03:17:47 +0000']]
	# my_titles = [['9919067856102836', 'November 1.'], ['9919067856102836', 'Best Of Sports Bhuja - March 29.'], ['9919067856202836', 'October 4.'], ['9919067856202836', 'Best Of Sports Bhuja - March 29.'], ['9919067856302836', 'September 20.'], ['9919067856302836', 'Best Of Sports Bhuja - March 29.'], ['9919067856402836', 'August 30.'], ['9919067856402836', 'Best Of Sports Bhuja - March 29.'], ['9919067856502836', 'August 16.'], ['9919067856502836', 'Best Of Sports Bhuja - March 29.'], ['9919067856602836', 'July 12.'], ['9919067856602836', 'Best Of Sports Bhuja - March 29.'], ['9919067856702836', 'July 5.'], ['9919067856702836', 'Best Of Sports Bhuja - March 29.'], ['9919067856802836', 'June 21.'], ['9919067856802836', 'Best Of Sports Bhuja - March 29.'], ['9919067856902836', 'June 14.'], ['9919067856902836', 'Best Of Sports Bhuja - March 29.'], ['9919067857002836', 'May 24.'], ['9919067857002836', 'Best Of Sports Bhuja - March 29.'], ['9919067857102836', 'May 17.'], ['9919067857102836', 'Best Of Sports Bhuja - March 29.'], ['9919067857202836', 'May 10.'], ['9919067857202836', 'Best Of Sports Bhuja - March 29.'], ['9919067857302836', 'May 3.'], ['9919067857302836', 'Best Of Sports Bhuja - March 29.'], ['9919067857402836', 'April 26.'], ['9919067857402836', 'Best Of Sports Bhuja - March 29.'], ['9919067857502836', 'April 12.'], ['9919067857502836', 'Best Of Sports Bhuja - March 29.'], ['9919067956002836', 'December 20.'], ['9919067956002836', 'Best Of Sports Bhuja - March 29.'], ['9919067956102836', 'December 13.'], ['9919067956102836', 'Best Of Sports Bhuja - March 29.'], ['9919067956202836', 'November 29.'], ['9919067956202836', 'Best Of Sports Bhuja - March 29.'], ['9919067956302836', 'December 6.'], ['9919067956302836', 'Best Of Sports Bhuja - March 29.'], ['9919067956402836', 'November 22.'], ['9919067956402836', 'Best Of Sports Bhuja - March 29.'], ['9919067956502836', 'November 15.'], ['9919067956502836', 'Best Of Sports Bhuja - March 29.'], ['9919067956602836', 'November 8.'], ['9919067956602836', 'Best of Sports Bhuja - November 8.'], ['9919067956702836', 'October 25.'], ['9919067956702836', 'Best Of Sports Bhuja - March 29.'], ['9919067956802836', 'October 11.'], ['9919067956802836', 'Best Of Sports Bhuja - March 29.'], ['9919067956902836', 'October 18.'], ['9919067956902836', 'Best Of Sports Bhuja - March 29.'], ['9919067957002836', 'September 27.'], ['9919067957002836', 'Best Of Sports Bhuja - March 29.'], ['9919067957102836', 'September 13.'], ['9919067957102836', 'Best Of Sports Bhuja - March 29.'], ['9919067957202836', 'September 6.'], ['9919067957202836', 'Best Of Sports Bhuja - March 29.'], ['9919067957302836', 'August 23.'], ['9919067957302836', 'Best Of Sports Bhuja - March 29.'], ['9919067957402836', 'August 9.'], ['9919067957402836', 'Best Of Sports Bhuja - March 29.'], ['9919067957502836', 'August 2.'], ['9919067957502836', 'Best Of Sports Bhuja - March 29.'], ['9919067957602836', 'July 26.'], ['9919067957602836', 'Best Of Sports Bhuja - March 29.'], ['9919067957702836', 'July 19.'], ['9919067957702836', 'Best Of Sports Bhuja - March 29.'], ['9919067957802836', 'June 28.'], ['9919067957802836', 'Best Of Sports Bhuja - March 29.'], ['9919067957902836', 'June 7.'], ['9919067957902836', 'Best Of Sports Bhuja - March 29.'], ['9919067958002836', 'May 31.'], ['9919067958002836', 'Best Of Sports Bhuja - March 29.'], ['9919067958102836', 'April 19.'], ['9919067958102836', 'Best Of Sports Bhuja - March 29.'], ['9919067958202836', 'April 5.'], ['9919067958202836', 'Best Of Sports Bhuja - March 29.'], ['9919067958302836', 'March 29.'], ['9919067958302836', 'Best Of Sports Bhuja - March 29.']]
	# mms_updated_list=[]
	# #mms_list = ['9919067856102836', '9919067856202836', '9919067856302836', '9919067856402836', '9919067856502836', '9919067856602836', '9919067856702836', '9919067856802836', '9919067856902836', '9919067857002836', '9919067857102836', '9919067857202836', '9919067857302836', '9919067857402836', '9919067857502836', '9919067956002836', '9919067956102836', '9919067956202836', '9919067956302836', '9919067956402836', '9919067956502836', '9919067956602836', '9919067956702836', '9919067956802836', '9919067956902836', '9919067957002836', '9919067957102836', '9919067957202836', '9919067957302836', '9919067957402836', '9919067957502836', '9919067957602836', '9919067957702836', '9919067957802836', '9919067957902836', '9919067958002836', '9919067958102836', '9919067958202836', '9919067958302836']
	# mms_list =  ['9919067856102836', '9919067856202836', '9919067856302836', '9919067856402836', '9919067856502836', '9919067856602836', '9919067856702836', '9919067856802836', '9919067856902836', '9919067857002836']

	# for st in sets:
	# 	offset = 0
	# 	offset_step = 99
	# 	title_list = []
	# 	for i in range(3):
			
	# 		print(i)
	# 		print(offset)
	# 		my_alma.get_set_members(st,{'limit':'100',"offset":str(offset)})
	# 		bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
	# 		for ind in range(len(bibs)):
	# 			my_alma.get_bib(bibs[ind])
	# 			my_rec = parse_xml_to_array(io.StringIO(my_alma.xml_response_data))[0]
	# 			if dateparser.parse(my_rec["245"]["a"].rstrip(".")):
	# 				print(my_rec)

	# 				f245a= my_rec["490"]["v"] + " - " + my_rec["245"]["a"]
	# 				try:
	# 					f490v = dateparser.parse(my_rec["245"]["a"].rstrip(".")).strftime('%B %d, %Y')
	# 				except:
	# 					f490v = my_rec["245"]["a"]
	# 					print(f490v)
	# 					print("?????????????????")
	# 				f830v = f490v + "."
	# 				my_rec["245"]["a"] = f245a
	# 				my_rec["490"]["v"] = f490v
	# 				my_rec["800"]["v"] = f830v
	# 				print(my_rec)
	# 				bib_data = start_xml +str(record_to_xml(my_rec)).replace("\\n", "\n").replace("\\", "")+end_xml
	# 				my_alma.update_bib(bibs[ind], bib_data)
	# 				print(bibs[ind]," - updated")
	# 				if bibs[ind] not in mms_updated_list:
	# 					mms_updated_list.append(bibs[ind])		
		




				# if my_rec["001"].data in mms_list:
				# 	print(my_rec)
				# 	f245a = " - ".join(my_rec["245"]["a"].split(" - ")[1:])
				# 	f490v = my_rec["245"]["a"].split(" - ")[0]
				# 	f830v = f490v + "."
				# 	my_rec["245"]["a"] = f245a
				# 	my_rec["490"]["v"] = f490v
				# 	my_rec["800"]["v"] = f830v
				# 	print(my_rec)
				# 	bib_data = start_xml +str(record_to_xml(my_rec)).replace("\\n", "\n").replace("\\", "")+end_xml
				# 	my_alma.update_bib(bibs[ind], bib_data)
				# 	print(bibs[ind]," - updated")
	# 			# 	if bibs[ind] not in mms_updated_list:
	# 					mms_updated_list.append(bibs[ind])
	# 		offset = offset+99



	# print(mms_updated_list)
	###############################################Adding 2 digits to 008#############################################################3
	# with open ("my_wrong_008_mms.txt") as f:
	# 	data = f.read()
	# 	for mms in data.split("\n")[:-1]:
	# 		my_alma.get_bib(mms)
	# 		my_rec = parse_xml_to_array(io.StringIO(my_alma.xml_response_data))[0]
	# 		f008 = my_rec["008"].data
	# 		print(f008)
	# 		if not "eng  " in f008:
	# 			new008 = f008.replace("eng##", "eng  ")
	# 			field = Field(tag = '008', data =new008)
	# 			my_rec.remove_fields("008")
	# 			my_rec.add_ordered_field(field) 
	# 			print(my_rec["245"]["a"])
	# 			# my_rec["245"]["a"] = my_rec["245"]["a"].replace("making.","making.  ")
	# 			print(my_rec)
	# 			bib_data = start_xml +str(record_to_xml(my_rec)).replace("\\n", "\n").replace("\\", "")+end_xml
	# 			print(bib_data)
	# 			my_alma.update_bib(mms, bib_data)
	# 			print(my_alma.xml_response_data)
	# 			print(my_alma.status_code)
	# 			print(mms," - updated")
	# 			quit()


	# with open ("missing_reps_mms.txt") as f:
	# 		data = f.read()
	# 		for mms in data.split("\n")[:-1]:
	# 			print (mms)
	# 			my_alma.get_bib(mms)
	# 			print(my_alma.xml_response_data)
	# # no_hold_mms = ['9919157198802836', '9919157198902836', '9919157199002836', '9919157199102836', '9919157199202836', '9919157199302836', '9919157199402836', '9919157199502836', '9919157199602836', '9919157199702836', '9919157199802836', '9919157298102836', '9919157298202836', '9919157298302836', '9919157298402836', '9919157298502836', '9919157298602836', '9919157298702836', '9919157298802836', '9919157298902836', '9919157299002836', '9919157299102836', '9919157299202836', '9919157299302836', '9919157299402836', '9919157299502836', '9919157299602836', '9919157299702836', '9919157299802836', '9919157299902836', '9919157300002836', '9919157300102836', '9919157300202836', '9919157300302836', '9919157300402836', '9919157300502836', '9919157300602836', '9919157300702836', '9919157300802836', '9919160138502836', '9919160138902836', '9919160139002836', '9919160139102836', '9919160139202836', '9919160139302836', '9919160139402836', '9919160139502836', '9919160139602836', '9919160139702836', '9919160139802836', '9919160139902836', '9919160140002836', '9919160140102836', '9919160140202836', '9919160239502836', '9919160239602836', '9919160239702836', '9919160239802836', '9919160239902836', '9919160240002836', '9919160240102836', '9919160240202836', '9919160240302836', '9919160240402836', '9919160241702836']
	# all_mms=['9919121706302836', '9919121706202836', '9919121706102836', '9919143372002836', '9919143271802836', '9919143371902836', '9919143371802836', '9919143271702836', '9919143271602836', '9919143271502836', '9919143371702836', '9919143371602836', '9919143271402836', '9919143371502836', '9919143271302836', '9919143271102836', '9919143271102836', '9919145776402836', '9919145776302836', '9919157202402836', '9919160241702836', '9919157202302836', '9919157202202836', '9919157202102836', '9919157202002836', '9919160241802836', '9919160142202836', '9919160142102836', '9919160142002836', '9919160241702836', '9919165395002836', '9919165495002836', '9919138743802836', '9919138743702836', '9919138843802836', '9919138743602836', '9919143271002836', '9919143371402836', '9919145776202836', '9919154183102836', '9919154282602836', '9919154183002836', '9919154282502836', '9919157201902836', '9919160142302836', '9919160242302836', '9919165394902836', '9919165494902836', '9919143371302836', '9919154282402836', '9919165394802836', '9919154182902836', '9919154182802836', '9919154282102836', '9919160242202836', '9919165494802836', '9919157201802836', '9919154189802836', '9919154189602836', '9919154189702836', '9919154189602836', '9919154189502836', '9919154189502836', '9919154189402836', '9919154189302836', '9919154189202836', '9919154288402836', '9919154288502836', '9919154288402836', '9919154288502836', '9919154288402836', '9919154288502836', '9919154288502836', '9919154288402836', '9919154288402836', '9919154288502836', '9919154288402836', '9919154288502836', '9919154288402836', '9919154288502836', '9919154288402836', '9919154288502836', '9919154288402836', '9919154288202836', '9919154288202836', '9919121607402836', '9919121606502836', '9919138843602836', '9919138843502836', '9919138743502836', '9919138843402836', '9919143371002836', '9919143370902836', '9919143270902836', '9919145776102836', '9919154182702836', '9919154282302836', '9919154182602836', '9919154282202836', '9919157201702836', '9919160242102836', '9919160242002836', '9919160241902836', '9919165394702836', '9919165494702836', '9919154286602836', '9919121607302836', '9919143370802836', '9919143370702836', '9919157201602836', '9919067938502836', '9919067841502836', '9919067841402836', '9919121707302836', '9919121607202836', '9919121707202836', '9919143370602836', '9919143270802836', '9919143370502836', '9919143370402836', '9919143370302836', '9919143370202836', '9919143370102836', '9919143370002836', '9919143369902836', '9919143270702836', '9919143270602836', '9919143270502836', '9919157201502836', '9919157201402836', '9919157201302836', '9919157302302836', '9919157201202836', '9919157201102836', '9919157201002836', '9919157200902836', '9919157302202836', '9919157302102836', '9919157200802836', '9919157200702836', '9919157302002836', '9919157301902836', '9919157301802836', '9919160141902836', '9919160241602836', '9919160141802836', '9919160141702836', '9919160141602836', '9919165394602836', '9919165494602836', '9919165494502836', '9919157200602836', '9919160241502836', '9919154186902836', '9919154186802836', '9919154286502836', '9919154186702836', '9919154186602836', '9919154186502836', '9919154186402836', '9919154186302836', '9919154286402836', '9919154286302836', '9919154186202836', '9919154186102836', '9919154286202836', '9919154286102836', '9919154186002836', '9919157200502836', '9919121606402836', '9919157301702836', '9919121606202836', '9919121606102836', '9919121607102836', '9919121606302836', '9919121606902836', '9919121707102836', '9919121705202836', '9919121707002836', '9919121706902836', '9919121606602836', '9919121606302836', '9919121705202836', '9919121706002836', '9919121705902836', '9919121705802836', '9919121705702836', '9919121705602836', '9919121705502836', '9919121705402836', '9919121705302836', '9919121606302836', '9919121705202836', '9919143368102836', '9919143267302836', '9919143368002836', '9919143369802836', '9919143269902836', '9919143367902836', '9919143269702836', '9919143269602836', '9919143269502836', '9919143267202836', '9919143267102836', '9919143269202836', '9919143367802836', '9919143267002836', '9919143369602836', '9919143369502836', '9919143269002836', '9919143268902836', '9919143367702836', '9919143266902836', '9919143367602836', '9919143367502836', '9919143367402836', '9919143268602836', '9919143268502836', '9919143367302836', '9919143369102836', '9919143369002836', '9919143267402836', '9919143368102836', '9919143267302836', '9919143368002836', '9919143367902836', '9919143267202836', '9919143267102836', '9919143367802836', '9919143267002836', '9919143367702836', '9919143266902836', '9919143367602836', '9919143367502836', '9919143367402836', '9919143367302836', '9919143267602836', '9919143267502836', '9919143267402836', '9919143368102836', '9919143267302836', '9919143368002836', '9919143367902836', '9919143267202836', '9919143267102836', '9919143367802836', '9919143267002836', '9919143367702836', '9919143266902836', '9919143367602836', '9919143367502836', '9919143367402836', '9919143367302836', '9919143367202836', '9919143365802836', '9919143367002836', '9919143366902836', '9919143266802836', '9919143366802836', '9919143266702836', '9919143266602836', '9919143366702836', '9919143366602836', '9919143365702836', '9919143266502836', '9919143366402836', '9919143366302836', '9919143366202836', '9919143366102836', '9919143366002836', '9919143365902836', '9919143365802836', '9919143365702836', '9919145876402836', '9919145876302836', '9919145876202836', '9919145876102836', '9919145876002836', '9919145776002836', '9919145775902836', '9919160140202836', '9919157301602836', '9919160140102836', '9919157200302836', '9919160140002836', '9919157301302836', '9919160139902836', '9919157301102836', '9919157301002836', '9919160240402836', '9919160240302836', '9919160240202836', '9919160139802836', '9919160240102836', '9919160240002836', '9919160139702836', '9919157199802836', '9919160139602836', '9919160139502836', '9919160139402836', '9919160239902836', '9919157300202836', '9919157300102836', '9919160139302836', '9919160139202836', '9919157199702836', '9919157199602836', '9919157299802836', '9919160239802836', '9919160239702836', '9919157299602836', '9919160239602836', '9919160139102836', '9919157299402836', '9919157299302836', '9919160139002836', '9919160138902836', '9919160239502836', '9919160141502836', '9919160241402836', '9919160141402836', '9919160141302836', '9919160141202836', '9919160141102836', '9919160241302836', '9919160241202836', '9919160241102836', '9919160241002836', '9919160240902836', '9919160141002836', '9919160140902836', '9919160140802836', '9919160140702836', '9919160140602836', '9919160140502836', '9919160140402836', '9919160240802836', '9919160240702836', '9919160240602836', '9919160140302836', '9919160240502836', '9919160140202836', '9919160140102836', '9919160140002836', '9919160139902836', '9919160240402836', '9919160240302836', '9919160240202836', '9919160139802836', '9919160240102836', '9919160240002836', '9919160139702836', '9919160139602836', '9919160139502836', '9919160139402836', '9919160239902836', '9919160139302836', '9919160139202836', '9919160239802836', '9919160239702836', '9919160239602836', '9919160139102836', '9919160139002836', '9919160138902836', '9919160239502836', '9919165494402836', '9919165494302836', '9919165394502836', '9919165394402836', '9919165394302836', '9919165394202836', '9919165494202836', '9919165494102836', '9919165494002836', '9919165493902836', '9919165394102836', '9919165394002836', '9919165393702836', '9919165393602836', '9919165493802836', '9919121705102836', '9919143365602836', '9919143266402836', '9919143365502836', '9919143365402836', '9919143365302836', '9919143365202836', '9919143266302836', '9919143365102836', '9919143365002836', '9919143266202836', '9919143364902836', '9919143266102836', '9919143364802836', '9919145875902836', '9919145875802836', '9919157299002836', '9919157199202836', '9919157298902836', '9919157298802836', '9919160138502836', '9919157199002836', '9919157298702836', '9919157298602836', '9919157198902836', '9919157298502836', '9919157298402836', '9919160138802836', '9919160239402836', '9919160239302836', '9919160138702836', '9919160138602836', '9919160239202836', '9919160239102836', '9919160138502836', '9919165393902836', '9919165393802836', '9919121705002836', '9919143265202836', '9919143363102836', '9919143363002836', '9919157298302836', '9919157298202836', '9919157198802836', '9919143362902836', '9919143362802836', '9919143265102836', '9919143265002836', '9919143362702836', '9919143362602836', '9919143362502836', '9919143362402836', '9919143264902836', '9919143264802836', '9919143362302836', '9919143264702836', '9919143362202836', '9919143362102836', '9919143264602836', '9919154183402836', '9919154282802836', '9919154282702836', '9919154183302836', '9919154183202836', '9919157298102836', '9919143362002836']
	# for el in no_hold_mms:
	# 	if not el in all_mms:
	# 		print(el)
	# with open ("no_holdings_mms.txt") as f:
	# 	data = f.read()
	# 	for mms in data.split("\n")[:-1]:
	# 		no_hold_mms.append(mms)
	# print(no_hold_mms)
	# print("here")
	# my_folder  =r"Y:\ndha\pre-deposit_prod\server_side_deposits\prod\ld_scheduled\oneoff_audio"
	# for serial_mms in os.listdir(my_folder):
	# 	if serial_mms in ["9919014268302836"]:
	# 		print("_____________________")
	# 		print(serial_mms)

	# 		for mis_mms in os.listdir(os.path.join(my_folder,serial_mms,"content")):
	# 			print(mis_mms)
	# 			my_alma.get_representations(mis_mms.split(".")[0])
	# 			print(my_alma.xml_response_data)
	# 			if mis_mms.split(".")[0] in no_hold_mms:
	# 				print("!!!!!!")



#Change Feedtalks <datafield tag="650" ind1="" ind2="7"> to <datafield tag="655" ind1="" ind2="7"> set  10530467590002836
# ###############################################Cult popture####################################################################
# 	mms_updated_list=[]
# 	bibs = ["9918975860202836"]
#my_alma = AlmaTools("prod")
# # sets = ["13345493180002836"]
# # sets = ["13345535710002836"]
# # sets = ["13345616790002836"]
# # #sets = ["13345446980002836"]
# # sets = ["13364936570002836"]#Lunch Money
# sets = ["12558695740002836"]#all podcasts MGR
# sets = ["13421281360002836"]
# # 	#Planning
# for st in sets:
# 	my_alma.get_set_members(st,{'limit':'100',"offset":str(0)})
# 	print(my_alma.xml_response_data)
# 	number_of_records = re.findall(r'total_record_count="(.*?)">', my_alma.xml_response_data)[0]
# 	offset = 0
# 	offset_step = 100
# 	title_list = []
# 	for i in range(int(number_of_records)//100+1):
# 			print(i)
# 			offset = offset+offset_step
# 			my_alma.get_set_members(st,{'limit':'100',"offset":str(offset)})
# 			#print(my_alma.xml_response_data)
# 			bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
# 			#print(bibs)
# 			for ind in range(len(bibs)):
# 				my_alma.get_bib(bibs[ind])
# 				#print(my_alma.xml_response_data)
# 				my_rec = parse_xml_to_array(io.StringIO(my_alma.xml_response_data))[0]
# 				print(my_rec["650"])
				# fields_650 = my_rec.get_fields('650')

				# for field in fields_650:
				# 	if 'Public policy.' in field.get_subfields('a'):
				# 		print("here")
				# 		my_rec.remove_field(field)

					
			
# # 				# f651 = Field(tag = '651', indicators = ["","0"], subfields = ['a', 'New Zealand', "x",'Social conditions.'])
# # 				# my_rec.add_ordered_field(f651)

				# bib_data = start_xml +str(record_to_xml(my_rec)).replace("\\n", "\n").replace("\\", "")+end_xml
				# my_alma.update_bib(bibs[ind], bib_data)
				# if my_alma.status_code in [200,202]:
				# 	print(bibs[ind]," - updated")
				# else:
				# 	print(bibs[ind]," - failed")

# all_bibs1 = []
# my_alma.get_set_members(set1,{'limit':'100',"offset":str(0)})
# bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
# all_bibs1 = list(bibs)

# my_alma.get_set_members(set1,{'limit':'100',"offset":str(100)})
# bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
# all_bibs1 = all_bibs1 +bibs
# print(len(all_bibs1))


# all_bibs2 = []

# my_alma.get_set_members(set2,{'limit':'100',"offset":str(0)})
# bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	
# all_bibs2 = list(bibs)

# my_alma.get_set_members(set2,{'limit':'100',"offset":str(100)})
# bibs = re.findall(r"<id>(.*?)</id>", my_alma.xml_response_data)	

# all_bibs2 = all_bibs2 +bibs
# print(len(all_bibs2))

# for el in all_bibs1:
# 	if not el in all_bibs2:
# 		print(el, " - record from updated set, which is not in new set")

# for el in all_bibs2:
# 	if not el in all_bibs1:
# 		print(el, " - record from new set, which is not in updated set")








if __name__ == '__main__':

	
	main()